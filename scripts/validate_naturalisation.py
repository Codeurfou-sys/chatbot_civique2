"""Contrôle la prise en compte des deux banques Naturalisation."""

from pathlib import Path
import re

from openpyxl import load_workbook


def values(path: Path, sheet_name: str, column: str) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    index = headers.index(column)
    result = [str(row[index] or "").strip() for row in rows if row[index]]
    workbook.close()
    return result


def normalize(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


module = Path("modules/05_preparer_examen.md").read_text(encoding="utf-8")
chatbot = Path("chat_bot.md").read_text(encoding="utf-8")
questions = values(Path("sources/BANQUE_OFFICIELLE_NATURALISATION.xlsx"), "Banque_NAT_Complete", "Question")
situations = values(Path("sources/MISES_EN_SITUATION_NATURALISATION.xlsx"), "Banque_MS_NAT_251", "Mise en situation")

missing_questions = [question for question in questions if normalize(question) not in normalize(module)]
knowledge_ids = re.findall(r"<!-- Source naturalisation : (NAT-[A-Z0-9-]+) -->", module)
situation_ids = re.findall(r"<!-- Source naturalisation : (MS-NAT-[A-Z0-9-]+) -->", module)
errors = []
if len(questions) != 251:
    errors.append(f"La banque officielle contient {len(questions)} questions au lieu de 251.")
if len(situations) != 251:
    errors.append(f"La banque de situations contient {len(situations)} lignes au lieu de 251.")
if missing_questions:
    errors.append(f"Questions non intégrées : {len(missing_questions)} (ex. {missing_questions[0]})")
if len(knowledge_ids) != 280 or len(set(knowledge_ids)) != 251:
    errors.append(f"Sources de connaissances : {len(knowledge_ids)} emplacements et {len(set(knowledge_ids))} identifiants uniques.")
if len(situation_ids) != 120 or len(set(situation_ids)) != 120:
    errors.append(f"Sources de situations : {len(situation_ids)} emplacements et {len(set(situation_ids))} identifiants uniques.")
for suffix in ("", "_VRAI", "_FAUX"):
    count = len(re.findall(rf"(?m)^## EXAM_NAT_V\d{{2}}_Q\d{{2}}{suffix}$", module))
    if count != 400:
        errors.append(f"Écrans Naturalisation {suffix or 'questions'} : {count} au lieu de 400.")
if module not in chatbot:
    errors.append("Le module 05 présent dans chat_bot.md n’est pas synchronisé.")
if errors:
    raise SystemExit("\n".join(errors))
print("OK — 251 questions utilisées, 120 situations issues de la banque, 10 examens Naturalisation synchronisés.")
