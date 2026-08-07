"""Intègre les banques Naturalisation dans les 10 examens blancs ChatMD.

Le script conserve les identifiants et la navigation des séries existantes,
mais remplace leurs 280 questions de connaissances et 120 mises en situation.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import argparse
import re

from openpyxl import load_workbook


CHAPTERS = {
    "T1_CH01": ("Les principes et valeurs de la République", "SCR_REV_T1_CH01_ACC"),
    "T1_CH02": ("La devise de la République française", "SCR_REV_T1_CH02_ACC"),
    "T1_CH03": ("Les symboles de la République française", "SCR_REV_T1_CH03_ACC"),
    "T1_CH04": ("La laïcité", "SCR_REV_T1_CH04_ACC"),
    "T1_CH05": ("La langue de la République", "SCR_REV_T1_CH05_ACC"),
    "T1_CH06": ("Le contrat d’engagement républicain", "SCR_REV_T1_CH06_ACC"),
    "T2_CH01": ("L’État de droit et la séparation des pouvoirs", "SCR_REV_T2_CH01_ACC"),
    "T2_CH02": ("La démocratie et le droit de vote", "SCR_REV_T2_CH02_ACC"),
    "T2_CH03": ("L’organisation et les institutions de la République", "SCR_REV_T2_CH03_ACC"),
    "T2_CH04": ("Les institutions européennes", "SCR_REV_T2_CH04_ACC"),
    "T3_CH01": ("Les droits fondamentaux", "SCR_REV_T3_CH01_ACC"),
    "T3_CH02": ("Les obligations et les devoirs", "SCR_REV_T3_CH02_ACC"),
    "T4_CH01": ("L’histoire de France", "SCR_REV_T4_CH01_ACC"),
    "T4_CH02": ("Les territoires et la géographie de la France", "SCR_REV_T4_CH02_ACC"),
    "T4_CH03": ("Le patrimoine et la culture française", "SCR_REV_T4_CH03_ACC"),
    "T5_CH01": ("Les démarches administratives", "SCR_REV_T5_CH01_ACC"),
    "T5_CH02": ("La santé", "SCR_REV_T5_CH02_ACC"),
    "T5_CH03": ("L’emploi", "SCR_REV_T5_CH03_ACC"),
    "T5_CH04": ("La parentalité et l’éducation", "SCR_REV_T5_CH04_ACC"),
}

KNOWLEDGE_PER_VARIANT = {1: 4, 2: 6, 3: 4, 4: 9, 5: 5}
SITUATIONS_PER_VARIANT = {1: 2, 2: 3, 3: 2, 4: 3, 5: 2}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def link_text(value: object) -> str:
    return clean(value).replace("]", r"\]")


def read_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    result = [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]
    workbook.close()
    return result


def chapter_key(row: dict[str, object]) -> str:
    theme = int(row["N° thématique"])
    label = clean(row["Chapitre"]).lower()
    explicit = re.search(r"chapitre\s+(\d+)", label)
    if explicit:
        return f"T{theme}_CH{int(explicit.group(1)):02d}"
    mappings = {
        1: [("symbole", 3), ("laïc", 4), ("langue", 5), ("engagement", 6), ("devise", 2)],
        2: [("union européenne", 4), ("relation", 4), ("élection", 2), ("démocr", 2),
            ("collectivité", 3), ("institution", 3), ("séparation", 1), ("justice", 1)],
        3: [("liberté", 1), ("droit fondamental", 1), ("texte fondateur", 1),
            ("devoir", 2), ("infraction", 2), ("justice", 2), ("citoyenneté", 2)],
        4: [("histoire", 1), ("géograph", 2), ("culture", 3), ("patrimoine", 3)],
        5: [("santé", 2), ("protection sociale", 2), ("emploi", 3), ("travail", 3),
            ("famil", 4), ("école", 4), ("parent", 4), ("logement", 1), ("société", 1)],
    }
    for needle, chapter in mappings[theme]:
        if needle in label:
            return f"T{theme}_CH{chapter:02d}"
    return f"T{theme}_CH01"


def distribute(rows: list[dict[str, object]], per_variant: dict[int, int]) -> list[list[dict[str, object]]]:
    by_theme: dict[int, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        by_theme[int(row["N° thématique"])].append(row)
    for values in by_theme.values():
        values.sort(key=lambda row: clean(row["ID"]))
    offsets = defaultdict(int)
    variants: list[list[dict[str, object]]] = []
    for variant in range(10):
        selected: list[dict[str, object]] = []
        for theme, count in per_variant.items():
            pool = by_theme[theme]
            for _ in range(count):
                selected.append(pool[offsets[theme] % len(pool)])
                offsets[theme] += 1
        # Rotation déterministe pour éviter dix séries organisées thème par thème.
        shift = (variant * 7) % len(selected)
        variants.append(selected[shift:] + selected[:shift])
    return variants


def replace_block(text: str, screen_id: str, body: str) -> str:
    pattern = re.compile(rf"(?ms)^## {re.escape(screen_id)}\s*$\n.*?(?=^## |\Z)")
    replacement = f"## {screen_id}\n\n{body.rstrip()}\n\n"
    text, count = pattern.subn(lambda _: replacement, text, count=1)
    if count != 1:
        raise RuntimeError(f"Écran introuvable ou dupliqué : {screen_id}")
    return text


def question_body(prefix: str, number: int, row: dict[str, object], situation: bool) -> str:
    timer = "?start=1" if number == 1 else ""
    variant_line = f"\n`@exam_variant = {int(prefix[10:12])}`\n" if number in (1, 29) else ""
    context = ""
    if situation:
        context = f"{clean(row['Mise en situation'])}\n\n"
        question = clean(row["Question posée"])
    else:
        question = clean(row["Question"])
    options = []
    correct = clean(row["Bonne réponse"]).upper()
    for letter in "ABCD":
        target = f"{prefix}_VRAI" if letter == correct else f"{prefix}_FAUX"
        options.append(f"{ord(letter) - 64}) [{link_text(row[f'Réponse {letter}'])}]({target})")
    return f"""`@err_{prefix[5:]} = 0`{variant_line}

<iframe
  src="https://codeurfou-sys.github.io/chatbot_civique2/minuteur-examen/{timer}"
  title="Minuteur de l’examen blanc"
  width="100%"
  height="94"
  loading="eager"
  style="border:0; border-radius:14px; background:#ffffff;"
></iframe>

### Question {number} sur 40

<!-- Source naturalisation : {clean(row['ID'])} -->

{context}**{question}**

{chr(10).join(options)}"""


def true_body(prefix: str, number: int, row: dict[str, object]) -> str:
    theme = int(row["N° thématique"])
    category = "connaissances" if number <= 28 else "situations"
    next_target = f"EXAM_NAT_V{prefix[10:12]}_Q{number + 1:02d}"
    label = "➡️ Question suivante"
    if number == 28:
        next_target = f"EXAM_NAT_V{prefix[10:12]}_PART2"
    elif number == 40:
        next_target = f"EXAM_NAT_V{prefix[10:12]}_RESULT"
        label = "📊 Accéder à mes résultats"
    return f"""`@exam_score = calc(@exam_score+1)`
`@exam_t{theme} = calc(@exam_t{theme}+1)`
`@exam_{category} = calc(@exam_{category}+1)`

1. [{label}]({next_target})"""


def false_body(prefix: str, number: int, row: dict[str, object]) -> str:
    key = chapter_key(row)
    next_target = f"EXAM_NAT_V{prefix[10:12]}_Q{number + 1:02d}"
    label = "➡️ Question suivante"
    if number == 28:
        next_target = f"EXAM_NAT_V{prefix[10:12]}_PART2"
    elif number == 40:
        next_target = f"EXAM_NAT_V{prefix[10:12]}_RESULT"
        label = "📊 Accéder à mes résultats"
    return f"""`@err_{prefix[5:]} = 1`

`@errchap_{key} = calc(@errchap_{key} + 1)`

1. [{label}]({next_target})"""


def correction(number: int, row: dict[str, object], situation: bool) -> str:
    question = clean(row["Question posée"] if situation else row["Question"])
    correct = clean(row["Bonne réponse"]).upper()
    answer = clean(row[f"Réponse {correct}"])
    explanation = clean(row["Feedback pédagogique"] if situation else row["Explication pédagogique"])
    tip = "" if situation else f"\n\n💡 {clean(row['Astuce mémoire'])}" if clean(row["Astuce mémoire"]) else ""
    return f"**{number}. {question}**  \n✅ {answer}\n\n{explanation}{tip}"


def recommendations() -> str:
    keys = list(CHAPTERS)
    parts = [
        "### 🎯 Conseils personnalisés",
        "",
        "Les recommandations ci-dessous sont calculées uniquement à partir des réponses incorrectes de cette série.",
        "",
    ]
    levels = [(">= 3", "🔴 Priorité forte", "Plusieurs erreurs ont été identifiées. Reprenez en priorité :"),
              ("== 2", "🟠 Priorité moyenne", "Ces chapitres méritent une révision ciblée :"),
              ("== 1", "🟡 Priorité faible", "Une erreur ponctuelle a été repérée. Vérifiez :")]
    for operator, title, intro in levels:
        condition = " || ".join(f"@errchap_{key} {operator}" for key in keys)
        parts.extend([f"`if {condition}`", f"#### {title}", "", intro, "", "`endif`"])
        for key, (label, target) in CHAPTERS.items():
            parts.extend([f"`if @errchap_{key} {operator}`", f"1. [📘 {label}]({target})", "`endif`"])
        parts.append("")
    none = " && ".join(f"@errchap_{key} == 0" for key in keys)
    any_error = " || ".join(f"@errchap_{key} >= 1" for key in keys)
    parts.extend([
        f"`if {none}`",
        "🟢 **Aucun chapitre à reprendre : toutes vos réponses sont correctes.**",
        "`endif`",
        "",
        f"`if {any_error}`",
        "Commencez par les priorités les plus fortes, puis réalisez un nouvel entraînement pour vérifier vos progrès.",
        "`endif`",
        "",
    ])
    return "\n".join(parts)


def integrate(module_path: Path, question_path: Path, situation_path: Path) -> None:
    questions = read_rows(question_path, "Banque_NAT_Complete")
    situations = read_rows(situation_path, "Banque_MS_NAT_251")
    by_id = {clean(row["ID"]): row for row in questions}
    situation_rows = []
    for row in situations:
        source = by_id[clean(row["ID question source"])]
        merged = {**source, **row, "N° thématique": source["N° thématique"], "Chapitre": source["Chapitre"]}
        situation_rows.append(merged)

    knowledge_variants = distribute(questions, KNOWLEDGE_PER_VARIANT)
    situation_variants = distribute(situation_rows, SITUATIONS_PER_VARIANT)
    text = module_path.read_text(encoding="utf-8")

    for variant in range(1, 11):
        rows = knowledge_variants[variant - 1] + situation_variants[variant - 1]
        for number, row in enumerate(rows, start=1):
            prefix = f"EXAM_NAT_V{variant:02d}_Q{number:02d}"
            is_situation = number >= 29
            text = replace_block(text, prefix, question_body(prefix, number, row, is_situation))
            text = replace_block(text, f"{prefix}_VRAI", true_body(prefix, number, row))
            text = replace_block(text, f"{prefix}_FAUX", false_body(prefix, number, row))
            cond_pattern = re.compile(
                rf"(?ms)(`if @err_NAT_V{variant:02d}_Q{number:02d} == 1`\n).*?(\n`endif`)"
            )
            replacement = rf"\1{correction(number, row, is_situation)}\2"
            text, count = cond_pattern.subn(lambda m, rep=replacement: re.sub(r"\\([12])", lambda x: m.group(int(x.group(1))), rep), text, count=1)
            if count != 1:
                raise RuntimeError(f"Corrigé introuvable : {prefix}")

        part1 = f"EXAM_NAT_V{variant:02d}_PART1"
        init_pattern = re.compile(rf"(?ms)(^## {part1}\s*$.*?`@exam_situations = 0`\n).*?(\n### 🧠 Partie 1 sur 2)")
        init_lines = "\n".join(f"`@errchap_{key} = 0`" for key in CHAPTERS)
        text, count = init_pattern.subn(lambda m: f"{m.group(1)}{init_lines}\n{m.group(2)}", text, count=1)
        if count != 1:
            raise RuntimeError(f"Initialisation introuvable : {part1}")

        totals = defaultdict(int)
        for row in rows:
            totals[int(row["N° thématique"])] += 1
        result_id = f"EXAM_NAT_V{variant:02d}_RESULT"
        result_pattern = re.compile(rf"(?ms)(^## {result_id}\s*$.*?#### Détail par thématique\n\n).*?(\n`if @exam_score >= 32`)")
        detail = "\n".join([
            f"- Thématique 1 — Principes et valeurs : **`@exam_t1` / {totals[1]}**",
            f"- Thématique 2 — Système institutionnel : **`@exam_t2` / {totals[2]}**",
            f"- Thématique 3 — Droits et devoirs : **`@exam_t3` / {totals[3]}**",
            f"- Thématique 4 — Histoire, géographie et culture : **`@exam_t4` / {totals[4]}**",
            f"- Thématique 5 — Vivre dans la société française : **`@exam_t5` / {totals[5]}**",
        ])
        text, count = result_pattern.subn(lambda m: f"{m.group(1)}{detail}\n{m.group(2)}", text, count=1)
        if count != 1:
            raise RuntimeError(f"Résultat introuvable : {result_id}")

        rec_pattern = re.compile(
            rf"(?ms)(^## {result_id}\s*$.*?)(### 🎯 Conseils personnalisés\n.*?)(?=1\. \[📘 Voir uniquement le corrigé)"
        )
        text, count = rec_pattern.subn(lambda m: f"{m.group(1)}{recommendations()}\n", text, count=1)
        if count != 1:
            raise RuntimeError(f"Recommandations introuvables : {result_id}")

    module_path.write_text(text, encoding="utf-8", newline="\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--module", default="modules/05_preparer_examen.md")
    parser.add_argument("--questions", default="sources/BANQUE_OFFICIELLE_NATURALISATION.xlsx")
    parser.add_argument("--situations", default="sources/MISES_EN_SITUATION_NATURALISATION.xlsx")
    args = parser.parse_args()
    integrate(Path(args.module), Path(args.questions), Path(args.situations))
    print("Naturalisation intégrée : 251 questions sources, 280 emplacements et 120 situations issues de la banque.")


if __name__ == "__main__":
    main()
