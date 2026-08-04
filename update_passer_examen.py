#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
update_passer_examen.py

Script intermédiaire du module 07 « Passer mon examen ».

Fonctions principales
---------------------
1. Charger les centres et sessions déjà présents dans le moteur Excel.
2. Tenter de récupérer les dernières sessions publiées sur le site FRATE.
3. Accepter également une source JSON locale pour sécuriser l'actualisation.
4. Supprimer toutes les dates passées.
5. Conserver uniquement les trois prochaines dates par centre.
6. Vérifier les liens Microsoft Forms.
7. Géocoder les centres via la Base Adresse Nationale (BAN).
8. Recalculer les centres les plus proches lorsqu'une localisation est fournie.
9. Mettre à jour sans risque les blocs 5 et 6 de l'onglet 07_PASSER_EXAMEN.
10. Générer les fichiers intermédiaires JSON utilisés par le futur export ChatMD.\n11. Ignorer proprement les cellules secondaires des plages fusionnées.

Dépendances
-----------
pip install openpyxl requests beautifulsoup4

Exemples
--------
Actualiser le classeur depuis le site FRATE :
    python update_passer_examen.py \
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx"

Utiliser un fichier JSON local :
    python update_passer_examen.py \
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx" \
        --sessions-json "sessions_frate.json"

Tester les centres proches de Strasbourg :
    python update_passer_examen.py \
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx" \
        --location "Strasbourg"

Ne pas modifier le classeur et générer uniquement les JSON :
    python update_passer_examen.py \
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx" \
        --dry-run
"""

from __future__ import annotations

import argparse
import copy
import dataclasses
import hashlib
import json
import logging
import math
import re
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


DEFAULT_FRATE_URL = "https://frateformation.net/formation/examen-civique/"
DEFAULT_BAN_URL = "https://api-adresse.data.gouv.fr/search/"
SHEET_MODULE = "07_PASSER_EXAMEN"
SHEET_SCREENS = "95_ECRANS_PASSER_EXAMEN"
MAX_SESSIONS_PER_CENTRE = 3
MAX_NEAREST_CENTRES = 3
REQUEST_TIMEOUT = 20
USER_AGENT = (
    "Mozilla/5.0 (compatible; CoachCiviqueUpdater/1.0; "
    "+https://frateformation.net/formation/examen-civique/)"
)


@dataclasses.dataclass
class Centre:
    code_region: str
    region: str
    code_centre: str
    ville: str
    departement: str
    forms_url: str
    screen_id: str
    ban_query: str
    active: bool = True
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    forms_ok: Optional[bool] = None
    forms_status: Optional[int] = None
    forms_checked_at: Optional[str] = None


@dataclasses.dataclass
class Session:
    code_centre: str
    region: str
    ville: str
    session_date: date
    forms_url: str
    source: str
    active: bool = True
    session_id: Optional[str] = None


@dataclasses.dataclass
class GeocodeResult:
    label: str
    city: str
    postcode: str
    context: str
    latitude: float
    longitude: float
    score: float


@dataclasses.dataclass
class NearestCentre:
    code_centre: str
    ville: str
    region: str
    departement: str
    distance_km: float
    forms_url: str
    sessions: list[str]


class Module07Error(RuntimeError):
    """Erreur fonctionnelle du module 07."""


def normalize_text(value: Any) -> str:
    """Normalise une chaîne pour faciliter les rapprochements."""
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def code_from_text(value: str) -> str:
    """Produit un identifiant stable en majuscules."""
    normalized = normalize_text(value)
    return re.sub(r"[^A-Z0-9]+", "_", normalized.upper()).strip("_")


def parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    return normalize_text(value) in {"oui", "yes", "true", "1", "actif", "active"}


def excel_or_text_to_date(value: Any) -> Optional[date]:
    """Convertit une valeur Excel, datetime ou texte en date."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            return None

    text = str(value).strip()
    formats = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%d/%m/%y",
        "%Y/%m/%d",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    match = re.search(
        r"\b(0?[1-9]|[12]\d|3[01])\s+"
        r"(janvier|février|fevrier|mars|avril|mai|juin|juillet|août|aout|"
        r"septembre|octobre|novembre|décembre|decembre)\s+(20\d{2})\b",
        normalize_text(text),
    )
    if match:
        months = {
            "janvier": 1,
            "fevrier": 2,
            "mars": 3,
            "avril": 4,
            "mai": 5,
            "juin": 6,
            "juillet": 7,
            "aout": 8,
            "septembre": 9,
            "octobre": 10,
            "novembre": 11,
            "decembre": 12,
        }
        return date(int(match.group(3)), months[match.group(2)], int(match.group(1)))
    return None


def date_to_iso(value: date) -> str:
    return value.isoformat()


def stable_session_id(code_centre: str, session_date: date) -> str:
    digest = hashlib.sha1(
        f"{code_centre}|{session_date.isoformat()}".encode("utf-8")
    ).hexdigest()[:8].upper()
    return f"SES_{code_centre}_{session_date:%Y%m%d}_{digest}"


def find_row_by_first_cell(ws, label: str) -> int:
    target = normalize_text(label)
    for row in range(1, ws.max_row + 1):
        if normalize_text(ws.cell(row=row, column=1).value) == target:
            return row
    raise Module07Error(
        f"Le libellé « {label} » est introuvable dans l'onglet {ws.title}."
    )


def find_block(
    ws,
    block_title: str,
    next_block_title: Optional[str] = None,
) -> tuple[int, int, int]:
    """
    Retourne (ligne_titre_bloc, ligne_entete, dernière_ligne_de_données_possible).
    """
    block_row = find_row_by_first_cell(ws, block_title)
    header_row = block_row + 1

    if next_block_title:
        end_row = find_row_by_first_cell(ws, next_block_title) - 1
    else:
        end_row = ws.max_row

    return block_row, header_row, end_row


def is_merged_placeholder(cell: Any) -> bool:
    """Retourne True pour une cellule secondaire appartenant à une fusion Excel."""
    return cell.__class__.__name__ == "MergedCell"


def safe_set_cell(ws, row: int, column: int, value: Any) -> bool:
    """
    Écrit une valeur sans provoquer d'AttributeError sur une cellule fusionnée.

    Les cellules secondaires d'une plage fusionnée sont ignorées. Le script ne
    déplace pas la valeur vers la cellule d'ancrage, afin de ne jamais écraser
    un titre ou un libellé de bloc.
    """
    cell = ws.cell(row=row, column=column)
    if is_merged_placeholder(cell):
        logging.debug(
            "Cellule fusionnée ignorée : %s!%s",
            ws.title,
            cell.coordinate,
        )
        return False
    cell.value = value
    return True


def safe_clear_cell(ws, row: int, column: int) -> bool:
    """Efface une cellule seulement si elle est réellement modifiable."""
    cell = ws.cell(row=row, column=column)
    if is_merged_placeholder(cell):
        return False
    cell.value = None
    return True


def clear_data_rows(ws, header_row: int, end_row: int) -> None:
    """
    Efface uniquement les valeurs situées sous l'en-tête d'un bloc.

    Les cellules secondaires des plages fusionnées sont laissées intactes.
    """
    for row in range(header_row + 1, end_row + 1):
        for col in range(1, ws.max_column + 1):
            safe_clear_cell(ws, row, col)


def load_centres_from_workbook(ws) -> list[Centre]:
    _, header_row, end_row = find_block(
        ws,
        "BLOC 5 — RÉGIONS ET CENTRES FRATE",
        "BLOC 6 — SESSIONS FUTURES",
    )

    centres: list[Centre] = []
    for row in range(header_row + 1, end_row + 1):
        code_centre = ws.cell(row=row, column=3).value
        if not code_centre:
            continue
        centres.append(
            Centre(
                code_region=str(ws.cell(row=row, column=1).value or "").strip(),
                region=str(ws.cell(row=row, column=2).value or "").strip(),
                code_centre=str(code_centre).strip(),
                ville=str(ws.cell(row=row, column=4).value or "").strip(),
                departement=str(ws.cell(row=row, column=5).value or "").strip(),
                forms_url=str(ws.cell(row=row, column=6).value or "").strip(),
                screen_id=str(ws.cell(row=row, column=8).value or "").strip(),
                ban_query=str(ws.cell(row=row, column=9).value or "").strip(),
                active=parse_bool(ws.cell(row=row, column=10).value),
            )
        )

    if not centres:
        raise Module07Error(
            "Aucun centre n'a été trouvé dans le bloc 5 de 07_PASSER_EXAMEN."
        )
    return centres


def load_sessions_from_workbook(ws) -> list[Session]:
    _, header_row, end_row = find_block(
        ws,
        "BLOC 6 — SESSIONS FUTURES",
        "BLOC 7 — CONFIGURATION BAN ET DISTANCES",
    )
    sessions: list[Session] = []
    for row in range(header_row + 1, end_row + 1):
        code_centre = ws.cell(row=row, column=2).value
        raw_date = ws.cell(row=row, column=5).value
        parsed_date = excel_or_text_to_date(raw_date)
        if not code_centre or not parsed_date:
            continue
        sessions.append(
            Session(
                session_id=str(ws.cell(row=row, column=1).value or "").strip() or None,
                code_centre=str(code_centre).strip(),
                region=str(ws.cell(row=row, column=3).value or "").strip(),
                ville=str(ws.cell(row=row, column=4).value or "").strip(),
                session_date=parsed_date,
                forms_url=str(ws.cell(row=row, column=7).value or "").strip(),
                source=str(ws.cell(row=row, column=8).value or "Classeur Excel").strip(),
                active=parse_bool(ws.cell(row=row, column=9).value),
            )
        )
    return sessions


def load_sessions_from_json(
    json_path: Path,
    centres_by_code: dict[str, Centre],
) -> list[Session]:
    raw = json.loads(json_path.read_text(encoding="utf-8"))
    records = raw.get("sessions", raw) if isinstance(raw, dict) else raw
    if not isinstance(records, list):
        raise Module07Error(
            "Le JSON doit contenir une liste ou un objet avec une clé « sessions »."
        )

    sessions: list[Session] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        code_centre = str(
            record.get("code_centre")
            or record.get("centre_code")
            or record.get("code")
            or ""
        ).strip()

        if not code_centre:
            ville_key = normalize_text(record.get("ville"))
            matches = [
                code
                for code, centre in centres_by_code.items()
                if normalize_text(centre.ville) == ville_key
            ]
            if len(matches) == 1:
                code_centre = matches[0]

        parsed_date = excel_or_text_to_date(
            record.get("date")
            or record.get("session_date")
            or record.get("date_session")
        )
        centre = centres_by_code.get(code_centre)
        if not code_centre or not parsed_date or not centre:
            logging.warning("Session JSON ignorée : %s", record)
            continue

        sessions.append(
            Session(
                code_centre=code_centre,
                region=str(record.get("region") or centre.region),
                ville=str(record.get("ville") or centre.ville),
                session_date=parsed_date,
                forms_url=str(
                    record.get("forms_url")
                    or record.get("lien_forms")
                    or centre.forms_url
                ),
                source=str(record.get("source") or f"JSON local : {json_path.name}"),
                active=parse_bool(record.get("actif"), True),
                session_id=str(record.get("id_session") or "").strip() or None,
            )
        )
    return sessions


def extract_dates_from_text(text: str) -> list[date]:
    """Extrait les formats de date courants d'un texte."""
    found: set[date] = set()

    numeric_patterns = (
        r"\b([0-3]?\d)[/-]([01]?\d)[/-](20\d{2})\b",
        r"\b(20\d{2})[/-]([01]?\d)[/-]([0-3]?\d)\b",
    )
    for match in re.finditer(numeric_patterns[0], text):
        try:
            found.add(date(int(match.group(3)), int(match.group(2)), int(match.group(1))))
        except ValueError:
            pass
    for match in re.finditer(numeric_patterns[1], text):
        try:
            found.add(date(int(match.group(1)), int(match.group(2)), int(match.group(3))))
        except ValueError:
            pass

    month_names = {
        "janvier": 1,
        "fevrier": 2,
        "février": 2,
        "mars": 3,
        "avril": 4,
        "mai": 5,
        "juin": 6,
        "juillet": 7,
        "aout": 8,
        "août": 8,
        "septembre": 9,
        "octobre": 10,
        "novembre": 11,
        "decembre": 12,
        "décembre": 12,
    }
    pattern = (
        r"\b([0-3]?\d)\s+("
        + "|".join(re.escape(month) for month in month_names)
        + r")\s+(20\d{2})\b"
    )
    for match in re.finditer(pattern, text, re.IGNORECASE):
        try:
            found.add(
                date(
                    int(match.group(3)),
                    month_names[match.group(2).lower()],
                    int(match.group(1)),
                )
            )
        except ValueError:
            pass
    return sorted(found)


def best_matching_centre(
    text: str,
    centres: Iterable[Centre],
) -> Optional[Centre]:
    normalized = normalize_text(text)
    scored: list[tuple[int, Centre]] = []

    for centre in centres:
        score = 0
        city = normalize_text(centre.ville)
        code = normalize_text(centre.code_centre.replace("_", " "))
        department = normalize_text(centre.departement)

        if city and city in normalized:
            score += 20 + len(city)
        if code and code in normalized:
            score += 10
        if department and re.search(rf"\b{re.escape(department)}\b", normalized):
            score += 3

        if score:
            scored.append((score, centre))

    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], item[1].ville))
    return scored[0][1]


def scrape_frate_sessions(
    url: str,
    centres: list[Centre],
    session: requests.Session,
) -> list[Session]:
    """
    Scraper volontairement tolérant.

    Il analyse les conteneurs HTML et rapproche chaque bloc d'une ville connue
    dans le moteur Excel. Les dates trouvées dans ce bloc sont associées au centre.
    """
    logging.info("Récupération des sessions FRATE : %s", url)
    response = session.get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    candidates: list[Any] = []

    selectors = (
        "article",
        "section",
        ".elementor-widget-container",
        ".elementor-tab-content",
        ".wp-block-group",
        ".accordion-item",
        ".et_pb_toggle",
        "li",
    )
    for selector in selectors:
        candidates.extend(soup.select(selector))

    if not candidates:
        candidates = [soup]

    extracted: list[Session] = []
    seen: set[tuple[str, date]] = set()

    for node in candidates:
        text = " ".join(node.stripped_strings)
        if len(text) < 10:
            continue
        centre = best_matching_centre(text, centres)
        if centre is None:
            continue
        dates = extract_dates_from_text(text)
        if not dates:
            continue

        forms_url = centre.forms_url
        for link in node.select("a[href]"):
            href = urljoin(url, link.get("href", "").strip())
            if "forms.office.com" in href or "forms.cloud.microsoft" in href:
                forms_url = href
                break

        for session_date in dates:
            key = (centre.code_centre, session_date)
            if key in seen:
                continue
            seen.add(key)
            extracted.append(
                Session(
                    code_centre=centre.code_centre,
                    region=centre.region,
                    ville=centre.ville,
                    session_date=session_date,
                    forms_url=forms_url,
                    source=f"Site FRATE : {url}",
                    active=True,
                )
            )

    logging.info("%d session(s) extraite(s) du site FRATE.", len(extracted))
    return extracted


def merge_sessions(*session_lists: Iterable[Session]) -> list[Session]:
    """
    Fusionne les sources.
    La dernière occurrence d'une clé centre/date remplace les précédentes.
    """
    merged: dict[tuple[str, date], Session] = {}
    for sessions in session_lists:
        for session in sessions:
            key = (session.code_centre, session.session_date)
            merged[key] = copy.deepcopy(session)
    return list(merged.values())


def filter_future_sessions(
    sessions: Iterable[Session],
    today: date,
    limit_per_centre: int = MAX_SESSIONS_PER_CENTRE,
) -> list[Session]:
    grouped: defaultdict[str, list[Session]] = defaultdict(list)

    for session in sessions:
        if not session.active:
            continue
        if session.session_date < today:
            continue
        grouped[session.code_centre].append(session)

    result: list[Session] = []
    for code_centre, values in grouped.items():
        values.sort(key=lambda item: item.session_date)
        unique_dates: set[date] = set()
        kept = 0
        for session in values:
            if session.session_date in unique_dates:
                continue
            unique_dates.add(session.session_date)
            session.session_id = stable_session_id(
                code_centre,
                session.session_date,
            )
            result.append(session)
            kept += 1
            if kept >= limit_per_centre:
                break

    result.sort(key=lambda item: (item.region, item.ville, item.session_date))
    return result


def validate_forms_url(url: str) -> bool:
    if not url:
        return False
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return False
    hostname = (parsed.hostname or "").casefold()
    return hostname in {
        "forms.office.com",
        "forms.cloud.microsoft",
        "forms.microsoft.com",
    }


def check_url(
    url: str,
    session: requests.Session,
) -> tuple[bool, Optional[int]]:
    if not validate_forms_url(url):
        return False, None

    try:
        response = session.head(
            url,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
        )
        status = response.status_code
        if status in {403, 405} or status >= 500:
            response = session.get(
                url,
                timeout=REQUEST_TIMEOUT,
                allow_redirects=True,
                stream=True,
            )
            status = response.status_code
        return 200 <= status < 400, status
    except requests.RequestException as exc:
        logging.warning("Lien Forms inaccessible : %s (%s)", url, exc)
        return False, None


def verify_forms_links(
    centres: list[Centre],
    session: requests.Session,
    skip_network: bool,
) -> None:
    cache: dict[str, tuple[bool, Optional[int]]] = {}
    checked_at = datetime.now().astimezone().isoformat(timespec="seconds")

    for centre in centres:
        if centre.forms_url in cache:
            ok, status = cache[centre.forms_url]
        elif skip_network:
            ok = validate_forms_url(centre.forms_url)
            status = None
            cache[centre.forms_url] = (ok, status)
        else:
            ok, status = check_url(centre.forms_url, session)
            cache[centre.forms_url] = (ok, status)
            time.sleep(0.15)

        centre.forms_ok = ok
        centre.forms_status = status
        centre.forms_checked_at = checked_at


def geocode(
    query: str,
    session: requests.Session,
    ban_url: str = DEFAULT_BAN_URL,
    limit: int = 5,
) -> list[GeocodeResult]:
    response = session.get(
        ban_url,
        params={"q": query, "limit": limit},
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()

    results: list[GeocodeResult] = []
    for feature in response.json().get("features", []):
        geometry = feature.get("geometry", {})
        coordinates = geometry.get("coordinates") or []
        properties = feature.get("properties", {})
        if len(coordinates) < 2:
            continue
        results.append(
            GeocodeResult(
                label=str(properties.get("label") or ""),
                city=str(properties.get("city") or properties.get("name") or ""),
                postcode=str(properties.get("postcode") or ""),
                context=str(properties.get("context") or ""),
                longitude=float(coordinates[0]),
                latitude=float(coordinates[1]),
                score=float(properties.get("score") or 0.0),
            )
        )
    return results


def load_geocode_cache(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
        return raw if isinstance(raw, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def save_geocode_cache(path: Path, cache: dict[str, dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def geocode_centres(
    centres: list[Centre],
    session: requests.Session,
    cache_path: Path,
    refresh: bool,
    skip_network: bool,
) -> None:
    cache = load_geocode_cache(cache_path)

    for centre in centres:
        cache_key = centre.code_centre
        cached = cache.get(cache_key)

        if cached and not refresh:
            centre.latitude = cached.get("latitude")
            centre.longitude = cached.get("longitude")
            continue

        if skip_network:
            logging.warning(
                "Coordonnées absentes pour %s et réseau désactivé.",
                centre.ville,
            )
            continue

        query = centre.ban_query or f"{centre.ville} {centre.departement}"
        try:
            results = geocode(query, session)
            if not results:
                logging.warning("Aucun géocodage BAN pour %s.", query)
                continue
            best = results[0]
            centre.latitude = best.latitude
            centre.longitude = best.longitude
            cache[cache_key] = {
                "code_centre": centre.code_centre,
                "ville": centre.ville,
                "query": query,
                "label": best.label,
                "latitude": best.latitude,
                "longitude": best.longitude,
                "score": best.score,
                "updated_at": datetime.now().astimezone().isoformat(
                    timespec="seconds"
                ),
            }
            time.sleep(0.15)
        except requests.RequestException as exc:
            logging.warning("Échec BAN pour %s : %s", query, exc)

    save_geocode_cache(cache_path, cache)


def haversine_km(
    lat1: float,
    lon1: float,
    lat2: float,
    lon2: float,
) -> float:
    radius = 6371.0088
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = (
        math.sin(delta_phi / 2) ** 2
        + math.cos(phi1)
        * math.cos(phi2)
        * math.sin(delta_lambda / 2) ** 2
    )
    return radius * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def calculate_nearest_centres(
    user_location: GeocodeResult,
    centres: list[Centre],
    sessions: list[Session],
    limit: int = MAX_NEAREST_CENTRES,
) -> list[NearestCentre]:
    sessions_by_centre: defaultdict[str, list[Session]] = defaultdict(list)
    for item in sessions:
        sessions_by_centre[item.code_centre].append(item)

    candidates: list[NearestCentre] = []
    for centre in centres:
        if not centre.active:
            continue
        if centre.latitude is None or centre.longitude is None:
            continue

        distance = haversine_km(
            user_location.latitude,
            user_location.longitude,
            centre.latitude,
            centre.longitude,
        )
        dates = sorted(
            {
                session.session_date.isoformat()
                for session in sessions_by_centre.get(centre.code_centre, [])
            }
        )
        candidates.append(
            NearestCentre(
                code_centre=centre.code_centre,
                ville=centre.ville,
                region=centre.region,
                departement=centre.departement,
                distance_km=round(distance, 1),
                forms_url=centre.forms_url,
                sessions=dates[:MAX_SESSIONS_PER_CENTRE],
            )
        )

    # Priorité aux centres avec au moins une session future.
    candidates.sort(
        key=lambda item: (
            0 if item.sessions else 1,
            item.distance_km,
            item.ville,
        )
    )
    return candidates[:limit]


def update_workbook_centres(
    ws,
    centres: list[Centre],
    sessions: list[Session],
) -> None:
    _, header_row, end_row = find_block(
        ws,
        "BLOC 5 — RÉGIONS ET CENTRES FRATE",
        "BLOC 6 — SESSIONS FUTURES",
    )
    clear_data_rows(ws, header_row, end_row)

    count_by_centre: defaultdict[str, int] = defaultdict(int)
    for session in sessions:
        count_by_centre[session.code_centre] += 1

    row = header_row + 1
    for centre in sorted(
        centres,
        key=lambda item: (item.region, item.ville),
    ):
        values = [
            centre.code_region,
            centre.region,
            centre.code_centre,
            centre.ville,
            centre.departement,
            centre.forms_url,
            count_by_centre.get(centre.code_centre, 0),
            centre.screen_id,
            centre.ban_query,
            "Oui" if centre.active else "Non",
        ]
        for col, value in enumerate(values, start=1):
            safe_set_cell(ws, row, col, value)
        row += 1


def update_workbook_sessions(
    ws,
    sessions: list[Session],
) -> None:
    _, header_row, end_row = find_block(
        ws,
        "BLOC 6 — SESSIONS FUTURES",
        "BLOC 7 — CONFIGURATION BAN ET DISTANCES",
    )
    clear_data_rows(ws, header_row, end_row)

    row = header_row + 1
    for item in sessions:
        values = [
            item.session_id or stable_session_id(
                item.code_centre,
                item.session_date,
            ),
            item.code_centre,
            item.region,
            item.ville,
            item.session_date,
            "À venir",
            item.forms_url,
            item.source,
            "Oui",
        ]
        for col, value in enumerate(values, start=1):
            safe_set_cell(ws, row, col, value)
        ws.cell(row=row, column=5).number_format = "dd/mm/yyyy"
        row += 1


def write_json_outputs(
    output_dir: Path,
    centres: list[Centre],
    sessions: list[Session],
    nearest: list[NearestCentre],
    location: Optional[GeocodeResult],
    source_summary: dict[str, Any],
) -> None:
    data_dir = output_dir / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    sessions_by_centre: defaultdict[str, list[Session]] = defaultdict(list)
    for item in sessions:
        sessions_by_centre[item.code_centre].append(item)

    centres_payload = []
    for centre in sorted(centres, key=lambda item: (item.region, item.ville)):
        centre_sessions = sessions_by_centre.get(centre.code_centre, [])
        centres_payload.append(
            {
                "code_region": centre.code_region,
                "region": centre.region,
                "code_centre": centre.code_centre,
                "ville": centre.ville,
                "departement": centre.departement,
                "forms_url": centre.forms_url,
                "forms_ok": centre.forms_ok,
                "forms_status": centre.forms_status,
                "forms_checked_at": centre.forms_checked_at,
                "screen_id": centre.screen_id,
                "latitude": centre.latitude,
                "longitude": centre.longitude,
                "sessions": [
                    {
                        "id_session": session.session_id,
                        "date": session.session_date.isoformat(),
                        "source": session.source,
                    }
                    for session in centre_sessions
                ],
                "actif": centre.active,
            }
        )

    sessions_payload = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "max_sessions_per_centre": MAX_SESSIONS_PER_CENTRE,
        "sources": source_summary,
        "centres": centres_payload,
    }
    (data_dir / "sessions.json").write_text(
        json.dumps(sessions_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    forms_payload = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "links": [
            {
                "code_centre": centre.code_centre,
                "ville": centre.ville,
                "url": centre.forms_url,
                "valid_syntax": validate_forms_url(centre.forms_url),
                "reachable": centre.forms_ok,
                "http_status": centre.forms_status,
                "checked_at": centre.forms_checked_at,
            }
            for centre in centres
        ],
    }
    (data_dir / "forms_validation.json").write_text(
        json.dumps(forms_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    if location is not None:
        nearest_payload = {
            "generated_at": datetime.now().astimezone().isoformat(
                timespec="seconds"
            ),
            "location": dataclasses.asdict(location),
            "nearest_centres": [
                dataclasses.asdict(item)
                for item in nearest
            ],
        }
        (data_dir / "nearest_centres.json").write_text(
            json.dumps(nearest_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )


def build_report(
    centres: list[Centre],
    sessions: list[Session],
    source_summary: dict[str, Any],
    nearest: list[NearestCentre],
) -> dict[str, Any]:
    links_ok = sum(1 for centre in centres if centre.forms_ok is True)
    links_bad = sum(1 for centre in centres if centre.forms_ok is False)
    geocoded = sum(
        1
        for centre in centres
        if centre.latitude is not None and centre.longitude is not None
    )
    sessions_by_centre: defaultdict[str, int] = defaultdict(int)
    for item in sessions:
        sessions_by_centre[item.code_centre] += 1

    return {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "centres_total": len(centres),
        "centres_geocodes": geocoded,
        "sessions_conservees": len(sessions),
        "centres_avec_session": len(sessions_by_centre),
        "limite_sessions_par_centre": MAX_SESSIONS_PER_CENTRE,
        "forms_valides": links_ok,
        "forms_invalides_ou_inaccessibles": links_bad,
        "sources": source_summary,
        "nearest_centres": [
            dataclasses.asdict(item)
            for item in nearest
        ],
    }


def configure_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(levelname)s — %(message)s",
    )


def create_http_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.5",
        }
    )
    return session


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Actualise le module 07 Passer mon examen avant l'export ChatMD."
        )
    )
    parser.add_argument(
        "--workbook",
        required=True,
        type=Path,
        help="Chemin du moteur Excel.",
    )
    parser.add_argument(
        "--output-workbook",
        type=Path,
        help=(
            "Classeur de sortie. Par défaut : <nom>_actualise.xlsx. "
            "Utiliser le même chemin que --workbook pour écraser le fichier."
        ),
    )
    parser.add_argument(
        "--sessions-json",
        type=Path,
        help="Source JSON locale facultative contenant les dernières sessions.",
    )
    parser.add_argument(
        "--frate-url",
        default=DEFAULT_FRATE_URL,
        help="Page FRATE contenant les centres et dates.",
    )
    parser.add_argument(
        "--ban-url",
        default=DEFAULT_BAN_URL,
        help="Point d'entrée de la Base Adresse Nationale.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("exports"),
        help="Dossier des fichiers JSON intermédiaires.",
    )
    parser.add_argument(
        "--location",
        help=(
            "Localisation facultative permettant de recalculer les trois "
            "centres les plus proches."
        ),
    )
    parser.add_argument(
        "--today",
        type=str,
        help="Date de référence au format AAAA-MM-JJ, utile pour les tests.",
    )
    parser.add_argument(
        "--no-scrape",
        action="store_true",
        help="Ne pas tenter de récupérer le site FRATE.",
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help="Désactiver les appels réseau BAN et Forms.",
    )
    parser.add_argument(
        "--refresh-geocodes",
        action="store_true",
        help="Ignorer le cache et recalculer les coordonnées des centres.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Ne pas enregistrer le classeur Excel.",
    )
    parser.add_argument(
        "--json-only",
        action="store_true",
        help=(
            "Générer les JSON et le rapport sans modifier les blocs 5 et 6 "
            "du classeur."
        ),
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Afficher les informations techniques.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    configure_logging(args.verbose)

    if not args.workbook.exists():
        logging.error("Classeur introuvable : %s", args.workbook)
        return 2

    if args.sessions_json and not args.sessions_json.exists():
        logging.error("JSON introuvable : %s", args.sessions_json)
        return 2

    try:
        today = (
            datetime.strptime(args.today, "%Y-%m-%d").date()
            if args.today
            else date.today()
        )
    except ValueError:
        logging.error("--today doit respecter le format AAAA-MM-JJ.")
        return 2

    output_workbook = args.output_workbook
    if output_workbook is None:
        output_workbook = args.workbook.with_name(
            f"{args.workbook.stem}_actualise{args.workbook.suffix}"
        )

    http = create_http_session()

    try:
        workbook = load_workbook(args.workbook)
    except Exception as exc:
        logging.exception("Impossible d'ouvrir le classeur : %s", exc)
        return 3

    if SHEET_MODULE not in workbook.sheetnames:
        logging.error("Onglet absent : %s", SHEET_MODULE)
        return 3
    if SHEET_SCREENS not in workbook.sheetnames:
        logging.warning(
            "Onglet écran absent : %s. Le script continuera sans le modifier.",
            SHEET_SCREENS,
        )

    ws = workbook[SHEET_MODULE]
    try:
        centres = load_centres_from_workbook(ws)
        workbook_sessions = load_sessions_from_workbook(ws)
    except Module07Error as exc:
        logging.error("%s", exc)
        return 4

    centres_by_code = {
        centre.code_centre: centre
        for centre in centres
    }

    source_summary: dict[str, Any] = {
        "excel": {
            "used": True,
            "sessions_loaded": len(workbook_sessions),
        },
        "json": {
            "used": False,
            "sessions_loaded": 0,
        },
        "frate": {
            "used": False,
            "sessions_loaded": 0,
            "url": args.frate_url,
        },
    }

    json_sessions: list[Session] = []
    if args.sessions_json:
        try:
            json_sessions = load_sessions_from_json(
                args.sessions_json,
                centres_by_code,
            )
            source_summary["json"] = {
                "used": True,
                "sessions_loaded": len(json_sessions),
                "path": str(args.sessions_json),
            }
            logging.info(
                "%d session(s) chargée(s) depuis %s.",
                len(json_sessions),
                args.sessions_json,
            )
        except (OSError, json.JSONDecodeError, Module07Error) as exc:
            logging.error("Lecture JSON impossible : %s", exc)
            return 4

    scraped_sessions: list[Session] = []
    if not args.no_scrape and not args.offline:
        try:
            scraped_sessions = scrape_frate_sessions(
                args.frate_url,
                centres,
                http,
            )
            source_summary["frate"]["used"] = True
            source_summary["frate"]["sessions_loaded"] = len(scraped_sessions)
        except requests.RequestException as exc:
            logging.warning(
                "Récupération FRATE impossible. "
                "Les sources Excel/JSON seront utilisées : %s",
                exc,
            )
        except Exception as exc:
            logging.warning(
                "Analyse de la page FRATE impossible : %s",
                exc,
            )

    # Ordre de priorité : Excel < scraping FRATE < JSON local.
    # Le JSON local est considéré comme une source contrôlée par l'administrateur.
    merged = merge_sessions(
        workbook_sessions,
        scraped_sessions,
        json_sessions,
    )
    filtered_sessions = filter_future_sessions(
        merged,
        today=today,
        limit_per_centre=MAX_SESSIONS_PER_CENTRE,
    )
    logging.info(
        "%d session(s) future(s) conservée(s), maximum %d par centre.",
        len(filtered_sessions),
        MAX_SESSIONS_PER_CENTRE,
    )

    # Harmoniser les URLs Forms avec le référentiel des centres.
    for item in filtered_sessions:
        centre = centres_by_code.get(item.code_centre)
        if centre and centre.forms_url:
            item.forms_url = centre.forms_url

    verify_forms_links(
        centres,
        http,
        skip_network=args.offline,
    )

    geocode_cache = args.output_dir / "data" / "centres_geocodes.json"
    geocode_centres(
        centres,
        http,
        cache_path=geocode_cache,
        refresh=args.refresh_geocodes,
        skip_network=args.offline,
    )

    location_result: Optional[GeocodeResult] = None
    nearest: list[NearestCentre] = []
    if args.location:
        if args.offline:
            logging.warning(
                "Le calcul de proximité est ignoré car --offline est actif."
            )
        else:
            try:
                location_candidates = geocode(
                    args.location,
                    http,
                    ban_url=args.ban_url,
                    limit=5,
                )
                if location_candidates:
                    location_result = location_candidates[0]
                    nearest = calculate_nearest_centres(
                        location_result,
                        centres,
                        filtered_sessions,
                    )
                else:
                    logging.warning(
                        "La BAN ne retourne aucun résultat pour : %s",
                        args.location,
                    )
            except requests.RequestException as exc:
                logging.warning("Recherche BAN impossible : %s", exc)

    if not args.json_only:
        update_workbook_centres(
            ws,
            centres,
            filtered_sessions,
        )
        update_workbook_sessions(
            ws,
            filtered_sessions,
        )
    else:
        logging.info(
            "Mode --json-only : les blocs Excel ne sont pas modifiés."
        )

    # Ne pas écrire dans B1 : dans le moteur Excel, la ligne de titre peut
    # contenir des cellules fusionnées. La date d'actualisation est enregistrée
    # dans les fichiers JSON et dans le rapport du module 07.

    write_json_outputs(
        args.output_dir,
        centres,
        filtered_sessions,
        nearest,
        location_result,
        source_summary,
    )

    report = build_report(
        centres,
        filtered_sessions,
        source_summary,
        nearest,
    )
    args.output_dir.mkdir(parents=True, exist_ok=True)
    report_path = args.output_dir / "module07_update_report.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    if not args.dry_run and not args.json_only:
        output_workbook.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_workbook)
        logging.info("Classeur actualisé : %s", output_workbook)
    elif args.json_only:
        logging.info(
            "Mode json-only : aucun classeur de sortie n'a été enregistré."
        )
    else:
        logging.info("Mode dry-run : le classeur n'a pas été enregistré.")

    logging.info("Rapport : %s", report_path)
    logging.info("Données sessions : %s", args.output_dir / "data" / "sessions.json")
    logging.info(
        "Validation Forms : %s",
        args.output_dir / "data" / "forms_validation.json",
    )

    if nearest:
        logging.info("Centres proches de %s :", args.location)
        for index, item in enumerate(nearest, start=1):
            logging.info(
                "%d. %s — %.1f km — %s",
                index,
                item.ville,
                item.distance_km,
                ", ".join(item.sessions) if item.sessions else "aucune date future",
            )

    return 0


if __name__ == "__main__":
    sys.exit(main())
