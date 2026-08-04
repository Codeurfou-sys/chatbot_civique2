#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
export_chatmd.py
================

Exporteur principal du moteur Excel « Coach Civique NovaFrate » vers ChatMD.

Le classeur Excel reste la source de vérité. Le script :

1. valide les identifiants d'écrans et la navigation ;
2. lit 00_ACCUEIL, 01_CONFIGURATION et les modules 02 à 10 ;
3. lit les catalogues d'écrans 90 à 98 ;
4. lit les transitions de 99_NAVIGATION ;
5. résout les contenus internes et les fichiers Markdown externes ;
6. génère un fichier principal ChatMD autonome et un fichier par module ;
7. génère les index CSV/JSON utiles à la FAQ, au glossaire,
   au module 07 et à la question libre ;
8. produit un rapport de validation et un manifeste d'export.

Dépendances
-----------
    pip install openpyxl

Exemple minimal
---------------
    python export_chatmd.py ^
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx" ^
        --output-dir "exports_chatmd"

Avec les fichiers Markdown sources
----------------------------------
    python export_chatmd.py ^
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx" ^
        --content-root "Ressources_MD" ^
        --output-dir "exports_chatmd"

Avec des URL GitHub Raw dans le fichier principal
--------------------------------------------------
    python export_chatmd.py ^
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx" ^
        --output-dir "exports_chatmd" ^
        --base-url "https://raw.githubusercontent.com/UTILISATEUR/DEPOT/main/chatmd"

Actualiser le module 07 avant l'export
--------------------------------------
    python export_chatmd.py ^
        --workbook "FICHIER_EXCEL_MOTEUR_CHAT_BOT.xlsx" ^
        --update-module07 ^
        --module07-updater "update_passer_examen.py" ^
        --output-dir "exports_chatmd"

Remarque importante
-------------------
ChatMD utilise notamment :
- des sections :  ## ID_ECRAN
- des boutons :   1. [Libellé](ID_ECRAN)
- des variables : @variable et @{variable}
- des calculs :   `@score = calc(@score+1)`
- des conditions : `if @variable == valeur`

Le fichier principal concatène tous les modules afin que ChatMD puisse fonctionner sans charger des inclusions relatives.

Le moteur Excel contient aussi des règles formulées en langage naturel.
Le script exporte directement les règles simples et signale dans le rapport
les conditions qui nécessitent encore une adaptation ChatMD explicite.
"""

from __future__ import annotations

import argparse
import csv
import dataclasses
import hashlib
import json
import logging
import re
import subprocess
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

DEFAULT_OUTPUT_DIR = Path("exports_chatmd")
DEFAULT_MAIN_FILE = "chat_bot.md"
DEFAULT_START_FILE = "start.md"
DEFAULT_ENCODING = "utf-8"

SCREEN_SHEETS = [
    "90_ECRANS_BILAN",
    "91_ECRANS_REVISIONS",
    "92_ECRANS_GLOSSAIRE",
    "93_ECRANS_PREPARER",
    "94_ECRANS_ENTRAINEMENT",
    "95_ECRANS_PASSER_EXAMEN",
    "96_ECRANS_CONSEILS",
    "97_ECRANS_FAQ",
    "98_ECRANS_QUESTION_LIBRE",
]

MODULE_SHEETS = [
    "02_BILAN",
    "03_REVISIONS",
    "04_GLOSSAIRE",
    "05_PREPARER_EXAMEN",
    "06_ENTRAINEMENT",
    "07_PASSER_EXAMEN",
    "08_CONSEILS",
    "09_FAQ",
    "10_QUESTION_LIBRE",
]

GLOBAL_SCREEN_IDS = {"START", "MENU_PRINCIPAL"}

ALLOWED_DYNAMIC_TARGET_PATTERNS = (
    re.compile(r"^\{[A-Za-z_][A-Za-z0-9_]*\}$"),
    re.compile(r"^@\{[A-Za-z_][A-Za-z0-9_]*\}$"),
    re.compile(r"^[A-Z0-9_]+\*$"),
    re.compile(r"^SCR_[A-Z0-9_]+_\*$"),
)

ACTIVE_VALUES = {"oui", "yes", "true", "1", "actif", "active"}
INACTIVE_VALUES = {"non", "no", "false", "0", "inactif", "inactive"}

MODULE_FILE_DEFAULTS = {
    "Accueil": "start.md",
    "Bilan": "02_bilan.md",
    "Révisions": "03_revisions.md",
    "Glossaire": "04_glossaire.md",
    "Préparer examen": "05_preparer_examen.md",
    "Préparer mon examen": "05_preparer_examen.md",
    "Entraînement": "06_entrainement.md",
    "Passer examen": "07_passer_examen.md",
    "Passer mon examen": "07_passer_examen.md",
    "Conseils": "08_conseils.md",
    "FAQ": "09_faq.md",
    "Question libre": "10_question_libre.md",
}

CONTENT_FIELD_LABELS = {
    "titre",
    "titre affiche",
    "titre affiché",
    "question",
    "reponse",
    "réponse",
    "contenu",
    "contenu markdown",
    "resume",
    "résumé",
    "objectif",
    "objectifs",
    "objectif general",
    "objectif général",
    "description",
    "texte",
    "message",
    "regle",
    "règle",
    "synthese",
    "synthèse",
    "points cles",
    "points clés",
    "glossaire",
    "ressources",
    "notes",
}


# ---------------------------------------------------------------------------
# Modèles de données
# ---------------------------------------------------------------------------

@dataclasses.dataclass
class ValidationIssue:
    level: str
    code: str
    message: str
    sheet: Optional[str] = None
    row: Optional[int] = None
    value: Optional[str] = None

    def to_dict(self) -> dict[str, Any]:
        return dataclasses.asdict(self)


@dataclasses.dataclass
class Screen:
    screen_id: str
    module: str
    subpath: str
    order: int
    screen_type: str
    title: str
    source_sheet: str
    content_ref: str
    variables: str
    input_expected: str
    business_rule: str
    version: str
    active: bool
    comment: str
    origin_sheet: str
    origin_row: int


@dataclasses.dataclass
class Navigation:
    nav_id: str
    module: str
    subpath: str
    source: str
    action_type: str
    label: str
    condition: str
    condition_value: str
    destination: str
    priority: int
    back_allowed: bool
    version: str
    active: bool
    origin_row: int


@dataclasses.dataclass
class WorkbookTable:
    sheet: str
    header_row: int
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclasses.dataclass
class ExportStats:
    generated_at: str
    workbook: str
    screens_total: int
    screens_exported: int
    navigation_total: int
    navigation_exported: int
    module_files: list[str]
    data_files: list[str]
    errors: int
    warnings: int


# ---------------------------------------------------------------------------
# Utilitaires
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value).strip())
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.casefold()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_header(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def parse_bool(value: Any, default: bool = True) -> bool:
    normalized = normalize_text(value)
    if not normalized:
        return default
    if normalized in ACTIVE_VALUES:
        return True
    if normalized in INACTIVE_VALUES:
        return False
    return default


def parse_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def is_url(value: str) -> bool:
    try:
        parsed = urlparse(value)
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)
    except Exception:
        return False


def is_dynamic_target(value: str) -> bool:
    if not value:
        return False
    if is_url(value):
        return True
    return any(pattern.match(value) for pattern in ALLOWED_DYNAMIC_TARGET_PATTERNS)


def chatmd_target(value: str) -> str:
    """
    Convertit {variable} vers @{variable}, syntaxe utilisable comme cible
    dynamique dans ChatMD.
    """
    value = safe_text(value)
    if re.fullmatch(r"\{[A-Za-z_][A-Za-z0-9_]*\}", value):
        return f"@{value}"
    return value


def slugify(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "module"


def clean_markdown_text(value: Any) -> str:
    text = safe_text(value)
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def markdown_escape_label(value: str) -> str:
    return (
        safe_text(value)
        .replace("\\", "\\\\")
        .replace("[", "\\[")
        .replace("]", "\\]")
    )


def atomic_write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(content, encoding=DEFAULT_ENCODING)
    temporary.replace(path)


def write_json(path: Path, payload: Any) -> None:
    atomic_write(
        path,
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
    )


def excel_value_to_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


# ---------------------------------------------------------------------------
# Lecture du classeur
# ---------------------------------------------------------------------------

class WorkbookReader:
    def __init__(self, workbook_path: Path, include_inactive: bool = False):
        self.workbook_path = workbook_path
        self.include_inactive = include_inactive
        self.wb = load_workbook(workbook_path, data_only=True, read_only=False)
        self.issues: list[ValidationIssue] = []
        self._content_index: dict[str, list[tuple[str, int, dict[str, Any]]]] = (
            defaultdict(list)
        )

    def add_issue(
        self,
        level: str,
        code: str,
        message: str,
        sheet: Optional[str] = None,
        row: Optional[int] = None,
        value: Optional[str] = None,
    ) -> None:
        self.issues.append(
            ValidationIssue(level, code, message, sheet, row, value)
        )

    def require_sheets(self) -> None:
        required = {
            "00_ACCUEIL",
            "01_CONFIGURATION",
            "99_NAVIGATION",
            *MODULE_SHEETS,
            *SCREEN_SHEETS,
        }
        for name in sorted(required):
            if name not in self.wb.sheetnames:
                self.add_issue(
                    "ERROR",
                    "MISSING_SHEET",
                    f"Onglet obligatoire absent : {name}",
                    sheet=name,
                )

    def iter_nonempty_rows(self, sheet_name: str) -> Iterator[tuple[int, list[Any]]]:
        ws = self.wb[sheet_name]
        for row_number in range(1, ws.max_row + 1):
            values = [
                ws.cell(row=row_number, column=column).value
                for column in range(1, ws.max_column + 1)
            ]
            if any(value not in (None, "") for value in values):
                yield row_number, values

    def find_header_rows(
        self,
        sheet_name: str,
        first_header_names: Iterable[str],
    ) -> list[int]:
        expected = {normalize_header(name) for name in first_header_names}
        rows: list[int] = []
        ws = self.wb[sheet_name]
        for row_number in range(1, ws.max_row + 1):
            first = normalize_header(ws.cell(row=row_number, column=1).value)
            if first in expected:
                rows.append(row_number)
        return rows

    def read_table_until_blank_or_block(
        self,
        sheet_name: str,
        header_row: int,
    ) -> WorkbookTable:
        ws = self.wb[sheet_name]
        headers = [
            normalize_header(ws.cell(row=header_row, column=column).value)
            for column in range(1, ws.max_column + 1)
        ]
        last_named_column = 0
        for index, header in enumerate(headers, start=1):
            if header:
                last_named_column = index
        headers = headers[:last_named_column]

        rows: list[dict[str, Any]] = []
        empty_count = 0
        for row_number in range(header_row + 1, ws.max_row + 1):
            first_value = ws.cell(row=row_number, column=1).value
            first_text = safe_text(first_value)

            if first_text.startswith("BLOC "):
                break

            values = [
                ws.cell(row=row_number, column=column).value
                for column in range(1, last_named_column + 1)
            ]
            if not any(value not in (None, "") for value in values):
                empty_count += 1
                if empty_count >= 1:
                    break
                continue

            empty_count = 0
            record = {
                headers[index]: values[index]
                for index in range(len(headers))
                if headers[index]
            }
            record["_row"] = row_number
            rows.append(record)

        return WorkbookTable(sheet_name, header_row, headers, rows)

    def index_all_content(self) -> None:
        """
        Indexe chaque ligne structurée de chaque onglet métier.

        Toute première colonne ressemblant à un identifiant devient une clé
        de résolution de contenu.
        """
        for sheet_name in ["00_ACCUEIL", *MODULE_SHEETS]:
            if sheet_name not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet_name]
            current_headers: list[str] = []

            for row_number in range(1, ws.max_row + 1):
                values = [
                    ws.cell(row=row_number, column=column).value
                    for column in range(1, ws.max_column + 1)
                ]
                first = safe_text(values[0]) if values else ""

                if first.startswith("BLOC "):
                    current_headers = []
                    continue

                if self._looks_like_header(values):
                    current_headers = [
                        normalize_header(value) for value in values
                    ]
                    continue

                if not current_headers or not first:
                    continue

                record: dict[str, Any] = {}
                for index, header in enumerate(current_headers):
                    if header and index < len(values):
                        record[header] = values[index]
                record["_row"] = row_number

                if self._looks_like_identifier(first):
                    self._content_index[first].append(
                        (sheet_name, row_number, record)
                    )

    @staticmethod
    def _looks_like_identifier(value: str) -> bool:
        if not value or len(value) > 120:
            return False
        if value.startswith("BLOC "):
            return False
        return bool(
            re.match(
                r"^(?:[A-Z]{1,12}[A-Z0-9_-]*|"
                r"SCR_[A-Z0-9_*{}-]+|"
                r"MSG_[A-Z0-9_-]+|"
                r"FAQ-\d+|"
                r"T\d(?:_CH\d+)?|"
                r"CONS_[A-Z0-9_-]+|"
                r"QLR_\d+|"
                r"INFO_[A-Z0-9_-]+|"
                r"SES_[A-Z0-9_-]+)$",
                value,
            )
        )

    @staticmethod
    def _looks_like_header(values: list[Any]) -> bool:
        normalized = [normalize_header(value) for value in values if value]
        if len(normalized) < 2:
            return False
        known = {
            "id_ecran",
            "id_navigation",
            "id_contenu",
            "id_faq",
            "id_chapitre",
            "id_theme",
            "id_message",
            "code_intention",
            "code_categorie",
            "cle",
            "parametre",
            "variable",
            "type_entite",
            "question_exemple",
            "code_region",
            "id_session",
        }
        return normalized[0] in known or len(set(normalized) & known) >= 1

    def resolve_content_records(
        self,
        content_ref: str,
    ) -> list[tuple[str, int, dict[str, Any]]]:
        return self._content_index.get(content_ref, [])

    def parse_screens(self) -> list[Screen]:
        screens: list[Screen] = []
        sheets = ["00_ACCUEIL", *SCREEN_SHEETS]

        for sheet_name in sheets:
            if sheet_name not in self.wb.sheetnames:
                continue

            header_rows = self.find_header_rows(sheet_name, ["ID écran"])
            for header_row in header_rows:
                table = self.read_table_until_blank_or_block(
                    sheet_name,
                    header_row,
                )
                for record in table.rows:
                    screen_id = safe_text(record.get("id_ecran"))
                    if not screen_id:
                        continue
                    active = parse_bool(record.get("actif"), True)
                    if not active and not self.include_inactive:
                        continue

                    screens.append(
                        Screen(
                            screen_id=screen_id,
                            module=safe_text(record.get("module")) or "Accueil",
                            subpath=safe_text(record.get("sous_parcours")),
                            order=parse_int(record.get("ordre"), 0),
                            screen_type=safe_text(record.get("type_d_ecran")),
                            title=safe_text(record.get("titre_affiche")),
                            source_sheet=safe_text(record.get("source_contenu")),
                            content_ref=safe_text(
                                record.get("contenu_reference")
                            ),
                            variables=safe_text(record.get("variables")),
                            input_expected=safe_text(
                                record.get("entree_utilisateur")
                            ),
                            business_rule=safe_text(
                                record.get("regle_metier")
                            ),
                            version=safe_text(record.get("version")) or "V1",
                            active=active,
                            comment=safe_text(
                                record.get("commentaire_technique")
                            ),
                            origin_sheet=sheet_name,
                            origin_row=int(record["_row"]),
                        )
                    )
        return screens

    def parse_navigation(self) -> list[Navigation]:
        if "99_NAVIGATION" not in self.wb.sheetnames:
            return []

        navigation: list[Navigation] = []
        for header_row in self.find_header_rows(
            "99_NAVIGATION",
            ["ID navigation"],
        ):
            table = self.read_table_until_blank_or_block(
                "99_NAVIGATION",
                header_row,
            )
            for record in table.rows:
                nav_id = safe_text(record.get("id_navigation"))
                if not nav_id:
                    continue
                active = parse_bool(record.get("actif"), True)
                if not active and not self.include_inactive:
                    continue

                navigation.append(
                    Navigation(
                        nav_id=nav_id,
                        module=safe_text(record.get("module")),
                        subpath=safe_text(record.get("sous_parcours")),
                        source=safe_text(record.get("ecran_source")),
                        action_type=safe_text(record.get("type_action")),
                        label=safe_text(record.get("libelle_affiche")),
                        condition=safe_text(record.get("condition")),
                        condition_value=safe_text(
                            record.get("valeur_condition")
                        ),
                        destination=safe_text(
                            record.get("ecran_destination")
                        ),
                        priority=parse_int(record.get("priorite"), 0),
                        back_allowed=parse_bool(
                            record.get("retour_autorise"),
                            True,
                        ),
                        version=safe_text(record.get("version")) or "V1",
                        active=active,
                        origin_row=int(record["_row"]),
                    )
                )
        return navigation

    def parse_configuration(self) -> dict[str, str]:
        config: dict[str, str] = {}
        if "01_CONFIGURATION" not in self.wb.sheetnames:
            return config

        ws = self.wb["01_CONFIGURATION"]
        for row_number in range(1, ws.max_row + 1):
            key = safe_text(ws.cell(row=row_number, column=1).value)
            value = ws.cell(row=row_number, column=2).value
            if not key or key.startswith("BLOC "):
                continue
            normalized_key = normalize_header(key).upper()
            if re.fullmatch(r"[A-Z][A-Z0-9_]+", normalized_key):
                config[normalized_key] = safe_text(value)
        return config

    def parse_module_registry(self) -> list[dict[str, Any]]:
        """
        Lit le bloc « REGISTRE DES MODULES » de 00_ACCUEIL.
        """
        if "00_ACCUEIL" not in self.wb.sheetnames:
            return []

        result: list[dict[str, Any]] = []
        for header_row in self.find_header_rows(
            "00_ACCUEIL",
            ["Code module"],
        ):
            table = self.read_table_until_blank_or_block(
                "00_ACCUEIL",
                header_row,
            )
            for record in table.rows:
                code = safe_text(record.get("code_module"))
                if not code:
                    continue
                active = parse_bool(record.get("actif"), True)
                if not active and not self.include_inactive:
                    continue
                result.append(
                    {
                        "code": code,
                        "label": safe_text(record.get("nom_affiche")),
                        "icon": safe_text(record.get("icone")),
                        "description": safe_text(
                            record.get("description_courte")
                        ),
                        "entry_screen": safe_text(
                            record.get("ecran_d_entree")
                        ),
                        "screen_sheet": safe_text(
                            record.get("onglet_ecrans")
                        ),
                        "order": parse_int(record.get("ordre"), 0),
                        "active": active,
                        "file": safe_text(
                            record.get("fichier_export_prevu")
                        ),
                    }
                )
        return sorted(result, key=lambda item: item["order"])


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

class Validator:
    def __init__(
        self,
        reader: WorkbookReader,
        screens: list[Screen],
        navigation: list[Navigation],
    ):
        self.reader = reader
        self.screens = screens
        self.navigation = navigation

    def run(self) -> list[ValidationIssue]:
        self.reader.require_sheets()
        self.validate_screen_ids()
        self.validate_navigation_ids()
        self.validate_navigation_sources()
        self.validate_navigation_destinations()
        self.validate_screen_content_refs()
        self.validate_variables()
        self.validate_module07_dates()
        return self.reader.issues

    def validate_screen_ids(self) -> None:
        occurrences: defaultdict[str, list[Screen]] = defaultdict(list)
        for screen in self.screens:
            occurrences[screen.screen_id].append(screen)

        for screen_id, values in occurrences.items():
            if len(values) > 1:
                locations = ", ".join(
                    f"{value.origin_sheet}!{value.origin_row}"
                    for value in values
                )
                self.reader.add_issue(
                    "ERROR",
                    "DUPLICATE_SCREEN_ID",
                    f"Écran déclaré plusieurs fois : {screen_id} ({locations})",
                    value=screen_id,
                )

    def validate_navigation_ids(self) -> None:
        occurrences: defaultdict[str, list[Navigation]] = defaultdict(list)
        for nav in self.navigation:
            occurrences[nav.nav_id].append(nav)

        for nav_id, values in occurrences.items():
            if len(values) > 1:
                rows = ", ".join(str(value.origin_row) for value in values)
                self.reader.add_issue(
                    "ERROR",
                    "DUPLICATE_NAV_ID",
                    f"Navigation déclarée plusieurs fois : {nav_id} "
                    f"(lignes {rows})",
                    sheet="99_NAVIGATION",
                    value=nav_id,
                )

    def validate_navigation_sources(self) -> None:
        screen_ids = {screen.screen_id for screen in self.screens}
        screen_ids |= GLOBAL_SCREEN_IDS

        for nav in self.navigation:
            if not nav.source:
                self.reader.add_issue(
                    "ERROR",
                    "EMPTY_NAV_SOURCE",
                    f"Source vide pour {nav.nav_id}",
                    sheet="99_NAVIGATION",
                    row=nav.origin_row,
                )
                continue
            if nav.source not in screen_ids and not is_dynamic_target(nav.source):
                self.reader.add_issue(
                    "ERROR",
                    "UNKNOWN_NAV_SOURCE",
                    f"Écran source inconnu : {nav.source}",
                    sheet="99_NAVIGATION",
                    row=nav.origin_row,
                    value=nav.source,
                )

    def validate_navigation_destinations(self) -> None:
        screen_ids = {screen.screen_id for screen in self.screens}
        screen_ids |= GLOBAL_SCREEN_IDS

        for nav in self.navigation:
            destination = nav.destination
            if not destination:
                self.reader.add_issue(
                    "ERROR",
                    "EMPTY_NAV_DESTINATION",
                    f"Destination vide pour {nav.nav_id}",
                    sheet="99_NAVIGATION",
                    row=nav.origin_row,
                )
                continue

            if destination in screen_ids:
                continue
            if is_dynamic_target(destination):
                self.reader.add_issue(
                    "INFO",
                    "DYNAMIC_NAV_DESTINATION",
                    f"Destination dynamique autorisée : {destination}",
                    sheet="99_NAVIGATION",
                    row=nav.origin_row,
                    value=destination,
                )
                continue

            self.reader.add_issue(
                "ERROR",
                "UNKNOWN_NAV_DESTINATION",
                f"Destination inconnue : {destination}",
                sheet="99_NAVIGATION",
                row=nav.origin_row,
                value=destination,
            )

    def validate_screen_content_refs(self) -> None:
        for screen in self.screens:
            reference = screen.content_ref
            if not reference:
                continue
            if reference.startswith("http"):
                continue
            if reference.startswith("{") and reference.endswith("}"):
                continue
            if reference.startswith("BLOC "):
                continue
            if reference.startswith("LISTE_"):
                continue
            if reference.startswith("CHAPITRE:"):
                continue
            if ".md" in reference.casefold():
                continue
            if screen.source_sheet in {"Markdown externe", "Fichier Markdown"}:
                continue

            records = self.reader.resolve_content_records(reference)
            if not records:
                self.reader.add_issue(
                    "WARNING",
                    "UNRESOLVED_CONTENT_REF",
                    f"Référence de contenu non résolue : {reference}",
                    sheet=screen.origin_sheet,
                    row=screen.origin_row,
                    value=reference,
                )

    def validate_variables(self) -> None:
        declared = self._declared_variables()
        used: set[str] = set()

        variable_pattern = re.compile(r"\{([A-Za-z_][A-Za-z0-9_]*)\}")
        for screen in self.screens:
            used |= set(variable_pattern.findall(screen.variables))
            used |= set(variable_pattern.findall(screen.content_ref))
            used |= set(variable_pattern.findall(screen.business_rule))
        for nav in self.navigation:
            used |= set(variable_pattern.findall(nav.condition_value))
            used |= set(variable_pattern.findall(nav.destination))

        for variable in sorted(used - declared):
            self.reader.add_issue(
                "WARNING",
                "UNDECLARED_VARIABLE",
                f"Variable utilisée mais non déclarée : {{{variable}}}",
                value=variable,
            )

    def _declared_variables(self) -> set[str]:
        declared: set[str] = set()
        pattern = re.compile(r"^\{([A-Za-z_][A-Za-z0-9_]*)\}$")

        for sheet_name in ["01_CONFIGURATION", *MODULE_SHEETS]:
            if sheet_name not in self.reader.wb.sheetnames:
                continue
            ws = self.reader.wb[sheet_name]
            for row in range(1, ws.max_row + 1):
                for column in range(1, min(ws.max_column, 5) + 1):
                    value = safe_text(
                        ws.cell(row=row, column=column).value
                    )
                    match = pattern.match(value)
                    if match:
                        declared.add(match.group(1))
        return declared

    def validate_module07_dates(self) -> None:
        if "07_PASSER_EXAMEN" not in self.reader.wb.sheetnames:
            return

        ws = self.reader.wb["07_PASSER_EXAMEN"]
        today = date.today()
        session_header_row = None

        for row in range(1, ws.max_row + 1):
            if normalize_header(ws.cell(row=row, column=1).value) == "id_session":
                session_header_row = row
                break

        if session_header_row is None:
            return

        counts: defaultdict[str, int] = defaultdict(int)
        for row in range(session_header_row + 1, ws.max_row + 1):
            first = safe_text(ws.cell(row=row, column=1).value)
            if first.startswith("BLOC "):
                break
            code_centre = safe_text(ws.cell(row=row, column=2).value)
            session_date = excel_value_to_date(
                ws.cell(row=row, column=5).value
            )
            if not code_centre or not session_date:
                continue
            counts[code_centre] += 1
            if session_date < today:
                self.reader.add_issue(
                    "ERROR",
                    "PAST_SESSION",
                    f"Session passée encore active : "
                    f"{code_centre} — {session_date.isoformat()}",
                    sheet="07_PASSER_EXAMEN",
                    row=row,
                )

        for code_centre, count in counts.items():
            if count > 3:
                self.reader.add_issue(
                    "WARNING",
                    "TOO_MANY_SESSIONS",
                    f"{code_centre} contient {count} sessions ; "
                    f"l'export ChatMD en affichera au maximum 3.",
                    sheet="07_PASSER_EXAMEN",
                    value=code_centre,
                )


# ---------------------------------------------------------------------------
# Résolution du contenu
# ---------------------------------------------------------------------------

class ContentResolver:
    def __init__(
        self,
        reader: WorkbookReader,
        content_root: Optional[Path],
        strict_content: bool = False,
    ):
        self.reader = reader
        self.content_root = content_root
        self.strict_content = strict_content

    def resolve(self, screen: Screen) -> str:
        reference = screen.content_ref

        if not reference:
            return self._fallback_from_screen(screen)

        if reference.startswith("CHAPITRE:"):
            return self._resolve_chapter_reference(reference)

        if ".md" in reference.casefold():
            return self._resolve_markdown_reference(reference, screen)

        records = self.reader.resolve_content_records(reference)
        if records:
            return self._render_record(records[0][2])

        if reference.startswith("LISTE_"):
            return (
                f"<!-- Liste dynamique : {reference}. "
                f"Les choix sont générés depuis 99_NAVIGATION. -->"
            )

        if reference.startswith("BLOC "):
            return f"<!-- Contenu généré depuis {reference}. -->"

        if reference.startswith("{") and reference.endswith("}"):
            variable = reference[1:-1]
            return f"`@{{{variable}}}`"

        if is_url(reference):
            return f"[Ouvrir la ressource]({reference})"

        return self._fallback_from_screen(screen)

    def _resolve_chapter_reference(self, reference: str) -> str:
        parts = reference.split(":")
        if len(parts) != 3:
            return f"<!-- Référence de chapitre invalide : {reference} -->"

        _, chapter_id, field = parts
        records = self.reader.resolve_content_records(chapter_id)
        if not records:
            return f"<!-- Chapitre introuvable : {chapter_id} -->"

        record = records[0][2]
        field_map = {
            "INTRO": ["introduction", "intro", "statut", "contenu"],
            "OBJECTIFS": ["objectifs", "objectif"],
            "SYNTHESE": ["synthese", "synthèse"],
            "VIGILANCE": [
                "points_de_vigilance",
                "vigilance",
                "ressources",
            ],
            "GLOSSAIRE": ["glossaire", "points_cles", "points_clés"],
        }
        candidates = field_map.get(field.upper(), [])
        for candidate in candidates:
            value = record.get(normalize_header(candidate))
            if value not in (None, "", "Non présent dans le cours source"):
                return clean_markdown_text(value)

        return (
            "<!-- Ce bloc n'est pas renseigné dans le moteur Excel "
            f"pour {chapter_id}:{field}. -->"
        )

    def _resolve_markdown_reference(
        self,
        reference: str,
        screen: Screen,
    ) -> str:
        if is_url(reference):
            return f"[Ouvrir le contenu source]({reference})"

        if self.content_root is None:
            return (
                f"<!-- Source Markdown attendue : {reference}. "
                f"Utiliser --content-root pour l'intégrer. -->"
            )

        candidates = [
            self.content_root / reference,
            self.content_root / Path(reference).name,
        ]
        for candidate in candidates:
            if candidate.exists() and candidate.is_file():
                return clean_markdown_text(
                    candidate.read_text(
                        encoding=DEFAULT_ENCODING,
                        errors="replace",
                    )
                )

        message = f"Fichier Markdown introuvable : {reference}"
        if self.strict_content:
            self.reader.add_issue(
                "ERROR",
                "MISSING_MARKDOWN_FILE",
                message,
                sheet=screen.origin_sheet,
                row=screen.origin_row,
            )
        else:
            self.reader.add_issue(
                "WARNING",
                "MISSING_MARKDOWN_FILE",
                message,
                sheet=screen.origin_sheet,
                row=screen.origin_row,
            )
        return f"<!-- {message} -->"

    def _render_record(self, record: dict[str, Any]) -> str:
        pieces: list[str] = []

        title_keys = [
            "titre",
            "titre_affiche",
            "question",
            "libelle",
            "nom_affiche",
        ]
        title = next(
            (
                clean_markdown_text(record.get(key))
                for key in title_keys
                if clean_markdown_text(record.get(key))
            ),
            "",
        )

        for key, value in record.items():
            if key.startswith("_") or value in (None, ""):
                continue
            normalized = normalize_header(key)
            if normalized in {
                "id",
                "id_contenu",
                "id_faq",
                "id_chapitre",
                "id_message",
                "code",
                "actif",
                "version",
                "source",
                "source_md",
                "id_ecran",
                "ecran",
                "ordre",
            }:
                continue
            if normalized in title_keys and clean_markdown_text(value) == title:
                continue

            text = clean_markdown_text(value)
            if not text or text == "Non présent dans le cours source":
                continue

            readable_label = key.replace("_", " ").strip().capitalize()
            if normalized in {
                "reponse",
                "contenu",
                "contenu_markdown",
                "texte",
                "message_regle",
                "message",
                "resume_contenu_de_reference",
            }:
                pieces.append(text)
            elif normalized in CONTENT_FIELD_LABELS:
                pieces.append(f"### {readable_label}\n\n{text}")
            elif len(text) > 120:
                pieces.append(f"### {readable_label}\n\n{text}")

        if not pieces:
            return title or "<!-- Contenu non renseigné. -->"
        return "\n\n".join(pieces)

    @staticmethod
    def _fallback_from_screen(screen: Screen) -> str:
        parts: list[str] = []
        if screen.business_rule:
            parts.append(f"<!-- Règle métier : {screen.business_rule} -->")
        if screen.input_expected and normalize_text(
            screen.input_expected
        ) not in {"aucune", "none"}:
            parts.append(
                f"**Réponse attendue :** {screen.input_expected}"
            )
        if not parts:
            parts.append("<!-- Contenu à compléter depuis le moteur Excel. -->")
        return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Compilation des conditions ChatMD
# ---------------------------------------------------------------------------

class ConditionCompiler:
    """
    Convertit seulement les conditions dont la traduction est certaine.

    Les conditions métier rédigées en français restent visibles dans des
    commentaires HTML et sont signalées dans validation_report.json.
    """

    SIMPLE_OPERATORS = {
        "==": "==",
        "!=": "!=",
        ">=": ">=",
        "<=": "<=",
        ">": ">",
        "<": "<",
    }

    def __init__(self, reader: WorkbookReader):
        self.reader = reader

    def compile(self, nav: Navigation) -> Optional[str]:
        condition = safe_text(nav.condition)
        value = safe_text(nav.condition_value)
        normalized = normalize_text(condition)

        if not condition or normalized in {"toujours", "always"}:
            return None

        # Condition déjà écrite dans la syntaxe ChatMD.
        if condition.lstrip().startswith("if "):
            return condition.strip()

        # Quelques formulations sûres.
        if normalized in {"reponse selectionnee", "réponse sélectionnée"}:
            self.reader.add_issue(
                "WARNING",
                "MANUAL_CONDITION_REQUIRED",
                f"La navigation {nav.nav_id} nécessite une affectation "
                f"de variable avant la destination ({value}).",
                sheet="99_NAVIGATION",
                row=nav.origin_row,
            )
            return None

        # Exemple accepté dans le moteur :
        # Variable @type_examen == CR
        match = re.search(
            r"(@[A-Za-z_][A-Za-z0-9_]*)\s*(==|!=|>=|<=|>|<)\s*(.+)",
            condition,
        )
        if match:
            variable, operator, expected = match.groups()
            return f"if {variable} {operator} {expected.strip()}"

        self.reader.add_issue(
            "INFO",
            "NATURAL_LANGUAGE_CONDITION",
            f"Condition conservée en commentaire : {condition}",
            sheet="99_NAVIGATION",
            row=nav.origin_row,
            value=condition,
        )
        return None


# ---------------------------------------------------------------------------
# Rendu ChatMD
# ---------------------------------------------------------------------------

class ChatMDRenderer:
    def __init__(
        self,
        reader: WorkbookReader,
        resolver: ContentResolver,
        screens: list[Screen],
        navigation: list[Navigation],
        base_url: Optional[str],
        variables_dynamic: bool,
        obfuscate: bool,
    ):
        self.reader = reader
        self.resolver = resolver
        self.screens = screens
        self.navigation = navigation
        self.base_url = base_url.rstrip("/") if base_url else None
        self.variables_dynamic = variables_dynamic
        self.obfuscate = obfuscate
        self.condition_compiler = ConditionCompiler(reader)

        self.navigation_by_source: defaultdict[str, list[Navigation]] = (
            defaultdict(list)
        )
        for item in navigation:
            self.navigation_by_source[item.source].append(item)
        for source in self.navigation_by_source:
            self.navigation_by_source[source].sort(
                key=lambda item: (item.priority, item.label, item.nav_id)
            )

    def render_screen(self, screen: Screen) -> str:
        lines: list[str] = [f"## {screen.screen_id}", ""]

        if screen.title:
            lines.extend([f"### {screen.title}", ""])

        if screen.variables:
            lines.extend(
                [
                    f"<!-- Variables : {screen.variables} -->",
                    "",
                ]
            )

        content = self.resolver.resolve(screen)
        if content:
            lines.extend([content, ""])

        buttons = self.navigation_by_source.get(screen.screen_id, [])
        if buttons:
            lines.extend(self._render_navigation(buttons))

        if screen.comment:
            lines.extend(
                [
                    "",
                    f"<!-- {screen.comment} -->",
                ]
            )

        return "\n".join(lines).rstrip() + "\n"

    def _render_navigation(
        self,
        navigation: list[Navigation],
    ) -> list[str]:
        lines: list[str] = []
        button_index = 1

        for nav in navigation:
            if nav.action_type.casefold() in {
                "automatique",
                "traitement invisible",
                "auto",
            }:
                lines.extend(self._render_automatic_transition(nav))
                continue

            condition_line = self.condition_compiler.compile(nav)
            if condition_line:
                lines.append(f"`{condition_line}`")

            if nav.condition and normalize_text(nav.condition) != "toujours":
                lines.append(
                    f"<!-- Condition métier : {nav.condition}"
                    + (
                        f" | Valeur : {nav.condition_value}"
                        if nav.condition_value
                        else ""
                    )
                    + " -->"
                )

            destination = chatmd_target(nav.destination)
            label = markdown_escape_label(nav.label or "Continuer")
            lines.append(
                f"{button_index}. [{label}]({destination})"
            )
            button_index += 1

        return lines

    def _render_automatic_transition(
        self,
        nav: Navigation,
    ) -> list[str]:
        destination = chatmd_target(nav.destination)
        lines = [
            f"<!-- Transition automatique {nav.nav_id} : "
            f"{nav.condition or 'Toujours'} → {destination} -->"
        ]

        condition_line = self.condition_compiler.compile(nav)
        if condition_line:
            lines.append(f"`{condition_line}`")

        # ChatMD n'a pas une syntaxe universelle de redirection automatique
        # documentée dans le moteur. On produit un bouton de continuation
        # masquable plutôt qu'un lien cassé.
        label = markdown_escape_label(nav.label or "Continuer")
        lines.append(f"1. [{label}]({destination})")
        return lines

    def render_module(
        self,
        module_name: str,
        module_screens: list[Screen],
    ) -> str:
        parts = [
            f"<!-- Module généré automatiquement : {module_name} -->",
            f"<!-- Date : {datetime.now().astimezone().isoformat(timespec='seconds')} -->",
            "",
        ]
        for screen in sorted(
            module_screens,
            key=lambda item: (
                item.subpath,
                item.order,
                item.screen_id,
            ),
        ):
            parts.append(self.render_screen(screen))
        return "\n".join(parts).rstrip() + "\n"

    def render_main_file(
        self,
        module_files: list[str],
        module_contents: list[tuple[str, str]],
        title: str,
        start_screen: str,
    ) -> str:
        """
        Génère un fichier ChatMD autonome.

        Les fichiers du dossier modules/ sont toujours conservés pour la
        maintenance, mais leur contenu est aussi concaténé dans chat_bot.md.
        Cette stratégie évite de dépendre du support des inclusions relatives
        par ChatMD ou du navigateur.
        """
        front_matter = [
            "---",
            f"obfuscate: {'true' if self.obfuscate else 'false'}",
            (
                "variablesDynamiques: true"
                if self.variables_dynamic
                else "variablesDynamiques: false"
            ),
            "---",
            "",
        ]

        body = [
            f"# {title}",
            "",
            f"1. [Commencer]({start_screen})",
            "",
        ]

        consolidated: list[str] = []
        seen_sections: set[str] = set()
        section_pattern = re.compile(r"^##\s+([A-Za-z0-9_*-]+)\s*$", re.MULTILINE)

        for filename, content in module_contents:
            cleaned = content.strip()
            if not cleaned:
                continue

            # Évite les doublons d'écrans dans le fichier final.
            sections = section_pattern.findall(cleaned)
            duplicate_sections = [
                section for section in sections if section in seen_sections
            ]
            if duplicate_sections:
                self.reader.add_issue(
                    "WARNING",
                    "DUPLICATE_SECTION_SKIPPED",
                    "Sections déjà présentes ignorées dans le fichier consolidé : "
                    + ", ".join(sorted(set(duplicate_sections))),
                    value=filename,
                )
                blocks = re.split(r"(?=^##\s+[A-Za-z0-9_*-]+\s*$)", cleaned, flags=re.MULTILINE)
                kept_blocks: list[str] = []
                for block in blocks:
                    match = re.match(r"^##\s+([A-Za-z0-9_*-]+)\s*$", block, re.MULTILINE)
                    if match and match.group(1) in seen_sections:
                        continue
                    kept_blocks.append(block)
                cleaned = "\n".join(block.strip() for block in kept_blocks if block.strip())

            for section in section_pattern.findall(cleaned):
                seen_sections.add(section)

            consolidated.extend(
                [
                    f"<!-- Début du fichier source : {filename} -->",
                    "",
                    cleaned,
                    "",
                    f"<!-- Fin du fichier source : {filename} -->",
                    "",
                ]
            )

        return "\n".join(front_matter + body + consolidated).rstrip() + "\n"

    def render_start_file(
        self,
        title: str,
        welcome: str,
        start_screen: str,
        home_screen: str,
    ) -> str:
        parts = [
            f"## {start_screen}",
            "",
            "<!-- Initialisation globale -->",
            f"1. [Ouvrir le Coach]({home_screen})",
            "",
            f"## {home_screen}",
            "",
            f"### {title}",
            "",
            welcome,
            "",
        ]
        parts.extend(
            self._render_navigation(
                self.navigation_by_source.get(home_screen, [])
            )
        )
        return "\n".join(parts).rstrip() + "\n"


# ---------------------------------------------------------------------------
# Exports de données
# ---------------------------------------------------------------------------

class DataExporter:
    def __init__(
        self,
        reader: WorkbookReader,
        output_data_dir: Path,
    ):
        self.reader = reader
        self.output_data_dir = output_data_dir
        self.generated: list[str] = []

    def export_all(self) -> list[str]:
        self.output_data_dir.mkdir(parents=True, exist_ok=True)
        self.export_faq()
        self.export_glossary()
        self.export_question_intents()
        self.export_sessions()
        self.export_variables()
        return self.generated

    def _write_csv(
        self,
        filename: str,
        rows: list[dict[str, Any]],
        preferred_fields: Optional[list[str]] = None,
    ) -> None:
        path = self.output_data_dir / filename
        if not rows:
            rows = [{"information": "Aucune donnée exportée"}]

        fields: list[str] = []
        if preferred_fields:
            fields.extend(
                field for field in preferred_fields
                if any(field in row for row in rows)
            )
        for row in rows:
            for key in row:
                if key.startswith("_"):
                    continue
                if key not in fields:
                    fields.append(key)

        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=fields,
                delimiter=";",
                extrasaction="ignore",
            )
            writer.writeheader()
            for row in rows:
                cleaned = {
                    key: self._serialise_value(row.get(key))
                    for key in fields
                }
                writer.writerow(cleaned)

        self.generated.append(str(path.relative_to(self.output_data_dir.parent)))

    @staticmethod
    def _serialise_value(value: Any) -> str:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return safe_text(value)

    def export_faq(self) -> None:
        rows = self._records_from_sheet_by_header(
            "09_FAQ",
            "id_faq",
        )
        self._write_csv(
            "faq.csv",
            rows,
            [
                "id_faq",
                "categorie",
                "ordre",
                "question",
                "reponse_markdown",
                "mots_cles",
                "id_ecran",
                "actif",
            ],
        )

    def export_glossary(self) -> None:
        rows: list[dict[str, Any]] = []
        if "04_GLOSSAIRE" in self.reader.wb.sheetnames:
            ws = self.reader.wb["04_GLOSSAIRE"]
            for header_row in self.reader.find_header_rows(
                "04_GLOSSAIRE",
                [
                    "ID",
                    "ID terme",
                    "ID_ENTREE",
                    "ID notion",
                    "Mot",
                    "Terme",
                ],
            ):
                table = self.reader.read_table_until_blank_or_block(
                    "04_GLOSSAIRE",
                    header_row,
                )
                rows.extend(table.rows)

        self._write_csv("glossaire.csv", rows)

    def export_question_intents(self) -> None:
        rows = self._records_from_sheet_by_header(
            "10_QUESTION_LIBRE",
            "id_regle",
        )
        self._write_csv(
            "intentions.csv",
            rows,
            [
                "id_regle",
                "intention",
                "type_correspondance",
                "expression_principale",
                "contexte_necessaire",
                "exclusion",
                "entite_attendue",
                "variable_produite",
                "ecran_si_complet",
                "priorite",
            ],
        )

        entities: list[dict[str, Any]] = []
        if "10_QUESTION_LIBRE" in self.reader.wb.sheetnames:
            for header_row in self.reader.find_header_rows(
                "10_QUESTION_LIBRE",
                ["Type entité"],
            ):
                table = self.reader.read_table_until_blank_or_block(
                    "10_QUESTION_LIBRE",
                    header_row,
                )
                entities.extend(table.rows)
        self._write_csv("entites.csv", entities)

    def export_sessions(self) -> None:
        sessions = self._records_from_sheet_by_header(
            "07_PASSER_EXAMEN",
            "id_session",
        )
        centres = self._records_from_sheet_by_header(
            "07_PASSER_EXAMEN",
            "code_region",
        )

        # Limite de sécurité : trois dates futures par centre.
        today = date.today()
        by_centre: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in sessions:
            code = safe_text(row.get("code_centre"))
            parsed = excel_value_to_date(row.get("date_session"))
            if not code or not parsed or parsed < today:
                continue
            row = dict(row)
            row["date_session"] = parsed.isoformat()
            by_centre[code].append(row)

        filtered: list[dict[str, Any]] = []
        for code, values in by_centre.items():
            values.sort(key=lambda item: safe_text(item.get("date_session")))
            filtered.extend(values[:3])

        payload = {
            "generated_at": datetime.now().astimezone().isoformat(
                timespec="seconds"
            ),
            "centres": centres,
            "sessions": filtered,
            "max_sessions_per_centre": 3,
        }
        path = self.output_data_dir / "sessions.json"
        write_json(path, payload)
        self.generated.append(
            str(path.relative_to(self.output_data_dir.parent))
        )

    def export_variables(self) -> None:
        variables: list[dict[str, Any]] = []
        for sheet_name in ["01_CONFIGURATION", *MODULE_SHEETS]:
            if sheet_name not in self.reader.wb.sheetnames:
                continue
            for header_row in self.reader.find_header_rows(
                sheet_name,
                ["Variable"],
            ):
                table = self.reader.read_table_until_blank_or_block(
                    sheet_name,
                    header_row,
                )
                for row in table.rows:
                    row = dict(row)
                    row["_source_sheet"] = sheet_name
                    variables.append(row)

        path = self.output_data_dir / "variables.json"
        write_json(
            path,
            {
                "generated_at": datetime.now()
                .astimezone()
                .isoformat(timespec="seconds"),
                "variables": variables,
            },
        )
        self.generated.append(
            str(path.relative_to(self.output_data_dir.parent))
        )

    def _records_from_sheet_by_header(
        self,
        sheet_name: str,
        first_header: str,
    ) -> list[dict[str, Any]]:
        if sheet_name not in self.reader.wb.sheetnames:
            return []
        rows: list[dict[str, Any]] = []
        for header_row in self.reader.find_header_rows(
            sheet_name,
            [first_header],
        ):
            table = self.reader.read_table_until_blank_or_block(
                sheet_name,
                header_row,
            )
            rows.extend(table.rows)
        return rows


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

class ExportApplication:
    def __init__(self, args: argparse.Namespace):
        self.args = args
        self.output_dir: Path = args.output_dir.resolve()
        self.modules_dir = self.output_dir / "modules"
        self.data_dir = self.output_dir / "data"
        self.reports_dir = self.output_dir / "reports"

    def run(self) -> int:
        self._prepare_output()

        workbook_path = self.args.workbook.resolve()
        if self.args.update_module07:
            workbook_path = self._run_module07_update(workbook_path)

        reader = WorkbookReader(
            workbook_path,
            include_inactive=self.args.include_inactive,
        )
        reader.index_all_content()

        screens = reader.parse_screens()
        navigation = reader.parse_navigation()
        config = reader.parse_configuration()
        registry = reader.parse_module_registry()

        validator = Validator(reader, screens, navigation)
        issues = validator.run()

        errors = [issue for issue in issues if issue.level == "ERROR"]
        warnings = [issue for issue in issues if issue.level == "WARNING"]

        self._write_validation_report(issues, screens, navigation)

        if errors and not self.args.force:
            logging.error(
                "%d erreur(s) bloquante(s). "
                "Consultez reports/validation_report.json.",
                len(errors),
            )
            return 5

        resolver = ContentResolver(
            reader,
            self.args.content_root.resolve()
            if self.args.content_root
            else None,
            strict_content=self.args.strict_content,
        )

        renderer = ChatMDRenderer(
            reader=reader,
            resolver=resolver,
            screens=screens,
            navigation=navigation,
            base_url=self.args.base_url,
            variables_dynamic=not self.args.disable_dynamic_variables,
            obfuscate=not self.args.disable_obfuscation,
        )

        module_file_map = self._build_module_file_map(
            screens,
            registry,
        )
        generated_module_files = self._write_module_files(
            renderer,
            screens,
            module_file_map,
        )

        start_filename = (
            config.get("START_MARKDOWN_FILE")
            or DEFAULT_START_FILE
        )
        start_path = self.modules_dir / Path(start_filename).name
        title = (
            config.get("PROJECT_NAME")
            or self._read_project_title(reader)
            or "Coach Civique NovaFrate"
        )
        start_screen = config.get("START_SCREEN_ID") or "START"
        home_screen = config.get("HOME_SCREEN_ID") or "MENU_PRINCIPAL"
        welcome = self._read_welcome_message(reader)

        atomic_write(
            start_path,
            renderer.render_start_file(
                title=title,
                welcome=welcome,
                start_screen=start_screen,
                home_screen=home_screen,
            ),
        )
        relative_start = str(start_path.relative_to(self.output_dir))
        if relative_start not in generated_module_files:
            generated_module_files.insert(0, relative_start)

        main_filename = (
            config.get("MAIN_MARKDOWN_FILE")
            or DEFAULT_MAIN_FILE
        )
        main_path = self.output_dir / Path(main_filename).name

        # Les includes doivent être relatifs au fichier principal ou absolus.
        include_paths = [
            str(Path(filename).as_posix())
            for filename in generated_module_files
        ]
        module_contents: list[tuple[str, str]] = []
        for relative_path in generated_module_files:
            module_path = self.output_dir / relative_path
            if not module_path.exists():
                reader.add_issue(
                    "ERROR",
                    "MISSING_GENERATED_MODULE",
                    f"Fichier de module généré introuvable : {relative_path}",
                    value=relative_path,
                )
                continue
            module_contents.append(
                (
                    relative_path,
                    module_path.read_text(
                        encoding=DEFAULT_ENCODING,
                        errors="replace",
                    ),
                )
            )

        atomic_write(
            main_path,
            renderer.render_main_file(
                module_files=include_paths,
                module_contents=module_contents,
                title=title,
                start_screen=start_screen,
            ),
        )

        data_exporter = DataExporter(reader, self.data_dir)
        generated_data_files = data_exporter.export_all()

        # Les résolutions de fichiers Markdown peuvent ajouter des problèmes.
        final_issues = reader.issues
        self._write_validation_report(
            final_issues,
            screens,
            navigation,
        )

        final_errors = sum(
            issue.level == "ERROR" for issue in final_issues
        )
        final_warnings = sum(
            issue.level == "WARNING" for issue in final_issues
        )

        stats = ExportStats(
            generated_at=datetime.now()
            .astimezone()
            .isoformat(timespec="seconds"),
            workbook=str(workbook_path),
            screens_total=len(screens),
            screens_exported=len(screens),
            navigation_total=len(navigation),
            navigation_exported=len(navigation),
            module_files=generated_module_files,
            data_files=generated_data_files,
            errors=final_errors,
            warnings=final_warnings,
        )
        write_json(
            self.output_dir / "manifest.json",
            dataclasses.asdict(stats),
        )
        self._write_readme(
            title,
            main_path.name,
            generated_module_files,
            generated_data_files,
            final_errors,
            final_warnings,
        )

        logging.info("Export terminé : %s", self.output_dir)
        logging.info("Fichier principal : %s", main_path)
        logging.info(
            "Écrans : %d | Navigation : %d | Erreurs : %d | Avertissements : %d",
            len(screens),
            len(navigation),
            final_errors,
            final_warnings,
        )
        return 0 if final_errors == 0 or self.args.force else 5

    def _prepare_output(self) -> None:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.modules_dir.mkdir(parents=True, exist_ok=True)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.reports_dir.mkdir(parents=True, exist_ok=True)

    def _run_module07_update(self, workbook_path: Path) -> Path:
        updater = self.args.module07_updater
        if updater is None:
            updater = Path("update_passer_examen.py")
        updater = updater.resolve()

        if not updater.exists():
            raise FileNotFoundError(
                f"Script du module 07 introuvable : {updater}"
            )

        updated_workbook = self.output_dir / (
            workbook_path.stem + "_module07_actualise.xlsx"
        )
        command = [
            sys.executable,
            str(updater),
            "--workbook",
            str(workbook_path),
            "--output-workbook",
            str(updated_workbook),
            "--output-dir",
            str(self.output_dir),
        ]
        if self.args.module07_sessions_json:
            command.extend(
                [
                    "--sessions-json",
                    str(self.args.module07_sessions_json.resolve()),
                ]
            )
        if self.args.offline:
            command.append("--offline")

        logging.info("Actualisation du module 07…")
        result = subprocess.run(command, check=False)
        if result.returncode != 0:
            raise RuntimeError(
                f"L'actualisation du module 07 a échoué "
                f"(code {result.returncode})."
            )
        return updated_workbook

    def _build_module_file_map(
        self,
        screens: list[Screen],
        registry: list[dict[str, Any]],
    ) -> dict[str, str]:
        mapping: dict[str, str] = {}

        for item in registry:
            label = item["label"]
            filename = item["file"]
            if filename:
                filename = Path(filename).name
            else:
                filename = (
                    MODULE_FILE_DEFAULTS.get(label)
                    or f"{item['order']:02d}_{slugify(label)}.md"
                )
            mapping[label] = filename

        for module_name in sorted({screen.module for screen in screens}):
            if module_name == "Accueil":
                continue
            if module_name not in mapping:
                mapping[module_name] = (
                    MODULE_FILE_DEFAULTS.get(module_name)
                    or f"{slugify(module_name)}.md"
                )
        return mapping

    def _write_module_files(
        self,
        renderer: ChatMDRenderer,
        screens: list[Screen],
        module_file_map: dict[str, str],
    ) -> list[str]:
        grouped: defaultdict[str, list[Screen]] = defaultdict(list)
        for screen in screens:
            if screen.screen_id in GLOBAL_SCREEN_IDS:
                continue
            if screen.module == "Accueil":
                grouped["Accueil"].append(screen)
            else:
                grouped[screen.module].append(screen)

        generated: list[str] = []
        for module_name, module_screens in sorted(grouped.items()):
            if module_name == "Accueil":
                filename = "00_accueil_complements.md"
            else:
                filename = module_file_map.get(
                    module_name,
                    f"{slugify(module_name)}.md",
                )
            path = self.modules_dir / Path(filename).name
            atomic_write(
                path,
                renderer.render_module(
                    module_name,
                    module_screens,
                ),
            )
            generated.append(str(path.relative_to(self.output_dir)))
        return generated

    def _write_validation_report(
        self,
        issues: list[ValidationIssue],
        screens: list[Screen],
        navigation: list[Navigation],
    ) -> None:
        payload = {
            "generated_at": datetime.now()
            .astimezone()
            .isoformat(timespec="seconds"),
            "summary": {
                "screens": len(screens),
                "navigation": len(navigation),
                "errors": sum(
                    issue.level == "ERROR" for issue in issues
                ),
                "warnings": sum(
                    issue.level == "WARNING" for issue in issues
                ),
                "information": sum(
                    issue.level == "INFO" for issue in issues
                ),
                "export_allowed": not any(
                    issue.level == "ERROR" for issue in issues
                ),
            },
            "issues": [issue.to_dict() for issue in issues],
        }
        write_json(
            self.reports_dir / "validation_report.json",
            payload,
        )

        lines = [
            "# Rapport de validation",
            "",
            f"- Écrans : **{len(screens)}**",
            f"- Transitions : **{len(navigation)}**",
            f"- Erreurs : **{payload['summary']['errors']}**",
            f"- Avertissements : **{payload['summary']['warnings']}**",
            "",
        ]
        for issue in issues:
            location = ""
            if issue.sheet:
                location = f" — `{issue.sheet}`"
                if issue.row:
                    location += f" ligne {issue.row}"
            lines.append(
                f"- **{issue.level} / {issue.code}** : "
                f"{issue.message}{location}"
            )
        atomic_write(
            self.reports_dir / "validation_report.md",
            "\n".join(lines) + "\n",
        )

    def _read_project_title(self, reader: WorkbookReader) -> str:
        if "00_ACCUEIL" not in reader.wb.sheetnames:
            return ""
        ws = reader.wb["00_ACCUEIL"]
        for row in range(1, ws.max_row + 1):
            if normalize_text(
                ws.cell(row=row, column=1).value
            ) == "nom du chatbot":
                return safe_text(ws.cell(row=row, column=2).value)
        return ""

    def _read_welcome_message(self, reader: WorkbookReader) -> str:
        records = reader.resolve_content_records("MSG_ACC_WELCOME")
        if records:
            record = records[0][2]
            for key in (
                "message_regle",
                "message",
                "texte",
                "contenu",
            ):
                value = clean_markdown_text(record.get(key))
                if value:
                    return value
        return (
            "Bienvenue dans le Coach Civique NovaFrate. "
            "Choisissez une rubrique pour commencer."
        )

    def _write_readme(
        self,
        title: str,
        main_filename: str,
        module_files: list[str],
        data_files: list[str],
        errors: int,
        warnings: int,
    ) -> None:
        lines = [
            f"# Export ChatMD — {title}",
            "",
            "## Fichier à ouvrir dans ChatMD",
            "",
            f"`{main_filename}`",
            "",
            "## Organisation",
            "",
            "- `modules/` : écrans générés depuis les onglets 00 et 90 à 98 ;",
            "- `data/` : CSV et JSON intermédiaires ;",
            "- `reports/` : contrôles de cohérence ;",
            "- `manifest.json` : résumé de l’export.",
            "",
            "## Contrôle avant publication",
            "",
            f"- Erreurs : **{errors}**",
            f"- Avertissements : **{warnings}**",
            "",
            "Consultez `reports/validation_report.md` avant de déposer "
            "les fichiers sur GitHub.",
            "",
            "## Fichiers de modules",
            "",
        ]
        lines.extend(f"- `{name}`" for name in module_files)
        lines.extend(["", "## Fichiers de données", ""])
        lines.extend(f"- `{name}`" for name in data_files)
        lines.extend(
            [
                "",
                "## Publication GitHub",
                "",
                "1. Déposer le contenu du dossier d’export dans le dépôt.",
                "2. Utiliser les URL `raw.githubusercontent.com` dans "
                "les inclusions ChatMD.",
                "3. Tester le menu principal, puis chaque module.",
                "4. Tester particulièrement le module 07 et la question libre.",
                "",
            ]
        )
        atomic_write(
            self.output_dir / "README_EXPORT.md",
            "\n".join(lines),
        )


# ---------------------------------------------------------------------------
# Interface en ligne de commande
# ---------------------------------------------------------------------------

def configure_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(levelname)s — %(message)s",
    )


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Transforme le moteur Excel du Coach Civique en fichiers ChatMD."
        )
    )
    parser.add_argument(
        "--workbook",
        required=True,
        type=Path,
        help="Classeur Excel moteur à exporter.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Dossier de sortie.",
    )
    parser.add_argument(
        "--content-root",
        type=Path,
        help=(
            "Dossier racine contenant les fichiers Markdown pédagogiques "
            "référencés dans Excel."
        ),
    )
    parser.add_argument(
        "--base-url",
        help=(
            "URL de base des fichiers publiés, par exemple une URL GitHub Raw. "
            "Sans cette option, les inclusions restent relatives."
        ),
    )
    parser.add_argument(
        "--include-inactive",
        action="store_true",
        help="Inclure les lignes dont Actif = Non.",
    )
    parser.add_argument(
        "--strict-content",
        action="store_true",
        help=(
            "Considérer tout fichier Markdown externe absent comme "
            "une erreur bloquante."
        ),
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Générer malgré les erreurs de validation.",
    )
    parser.add_argument(
        "--disable-obfuscation",
        action="store_true",
        help="Générer `obfuscate: false` dans le fichier principal.",
    )
    parser.add_argument(
        "--disable-dynamic-variables",
        action="store_true",
        help="Générer `variablesDynamiques: false`.",
    )
    parser.add_argument(
        "--update-module07",
        action="store_true",
        help=(
            "Exécuter update_passer_examen.py avant l'export."
        ),
    )
    parser.add_argument(
        "--module07-updater",
        type=Path,
        help="Chemin de update_passer_examen.py.",
    )
    parser.add_argument(
        "--module07-sessions-json",
        type=Path,
        help="JSON local facultatif transmis au script du module 07.",
    )
    parser.add_argument(
        "--offline",
        action="store_true",
        help=(
            "Transmettre le mode hors ligne au script du module 07."
        ),
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Afficher les détails techniques.",
    )
    return parser


def main() -> int:
    parser = build_argument_parser()
    args = parser.parse_args()
    configure_logging(args.verbose)

    if not args.workbook.exists():
        logging.error("Classeur introuvable : %s", args.workbook)
        return 2

    try:
        application = ExportApplication(args)
        return application.run()
    except KeyboardInterrupt:
        logging.error("Export interrompu.")
        return 130
    except Exception as exc:
        logging.exception("Échec de l'export : %s", exc)
        return 10


if __name__ == "__main__":
    sys.exit(main())
